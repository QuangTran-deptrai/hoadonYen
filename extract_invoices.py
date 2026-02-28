import re
import os
import sys
import pdfplumber
import pandas as pd
import openpyxl
import ast  # Added for parsing dict strings
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# OCR imports (optional - for scanned PDFs)
OCR_AVAILABLE = False
POPPLER_PATH = None  # Use system default on Linux
try:
    import pytesseract
    from pdf2image import convert_from_path
    from PIL import Image
    OCR_AVAILABLE = True
    # Configure Tesseract path (Windows only, Linux uses system default)
    import platform
    if platform.system() == "Windows":
        # Check for bundled env first
        if getattr(sys, 'frozen', False):
             base_path = sys._MEIPASS
        else:
             base_path = os.path.dirname(os.path.abspath(__file__))
             
        tesseract_paths = [
            os.path.join(base_path, 'tesseract', 'tesseract.exe'), # Bundled
            r"C:\Program Files\Tesseract-OCR\tesseract.exe",
            r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
            os.path.join(os.getenv('LOCALAPPDATA', ''), 'Tesseract-OCR', 'tesseract.exe')
        ]
        tess_found = False
        for p in tesseract_paths:
            if os.path.exists(p):
                pytesseract.pytesseract.tesseract_cmd = p
                tess_found = True
                break
        
        # Determine Poppler path
        bundled_poppler = os.path.join(base_path, "poppler-24.08.0", "Library", "bin")
        if os.path.exists(bundled_poppler):
            POPPLER_PATH = bundled_poppler
        else:
            # Fallback to hardcoded dev path
            possible_path = r"d:\hoadon\poppler-24.08.0\Library\bin"
            if os.path.exists(possible_path):
                POPPLER_PATH = possible_path
except ImportError:
    pass  # OCR not available, will skip scanned PDFs

# Category mapping based on extracted services
CATEGORY_KEYWORDS = {
    "Dịch vụ ăn uống": [
        "món", "lẩu", "gà", "bò", "heo", "cá", "cua", "mực", "tôm", "ghẹ", "sò", 
        "gỏi", "xào", "nướng", "chiên", "hấp", "hầm", "quay", "cơm", "xôi", "soup",
        "trà", "cà phê", "nước", "coca", "matcha", "oolong", "trái cây", "bánh",
        "đậu", "trứng", "lươn", "hàu", "khô mực", "khăn lạnh", "hủ tiếu", "baba",
        "bồ câu", "chả", "dừa", "khoáng", "suối", "sả", "rượu", "coffee", "cafe",
        "katinat", "highlands", "starbucks", "phúc long", "trung nguyên", "golden gate",
        "pizza", "kfc", "lotteria", "jollibee", "mcdonald", "domino","Xa lach rau mam", "buffet"
    ],
    "Dịch vụ phòng nghỉ": [
        "thuê phòng", "phòng số", "cho thuê phòng", "phòng họp", "meeting room",
        "hotel", "khách sạn", "homestay", "resort"
    ],
    "Hoa tươi": [
        "hoa tươi", "hoa", "bó hoa", "lãng hoa", "kệ hoa"
    ],
    "Thẻ cào điện thoại": [
        "cước", "di động", "thẻ cào", "sim", "điện thoại", "internet", "mạng", "mệnh giá",
        "THE CAO MENH GIA", "viettel", "mobifone", "vinaphone"
    ],
    "Xăng xe": [
        "xăng", "dầu", "diesel", "ron95", "ron92", "e5", "do 0.05s", "petrolimex"
    ],
    "Quà tặng": [
        "quà", "gift", "quà tặng", "tặng phẩm"
    ]
}

# F&B brand names for seller-based classification
FB_BRANDS = ["KATINAT", "HIGHLANDS", "STARBUCKS", "PHÚC LONG", "COFFEE HOUSE", 
             "TRUNG NGUYÊN", "GOLDEN GATE", "PIZZA", "KFC", "LOTTERIA", 
             "JOLLIBEE", "MCDONALD", "DOMINO"]

def classify_content(services_text, seller_name=""):
    """Classify services into categories using keyword matching with word boundaries."""
    
    # Priority: Check Seller Name for known F&B brands
    if seller_name:
        seller_upper = seller_name.upper()
        if any(brand in seller_upper for brand in FB_BRANDS):
            return "Dịch vụ ăn uống"

    if not services_text:
        return "Khác"
    text_lower = services_text.lower()
    
    scores = {}
    for category, keywords in CATEGORY_KEYWORDS.items():
        score = 0
        for kw in keywords:
            # Use regex word boundary to avoid partial matches (e.g. "thấp" matching "hấp")
            # BUT for "Xăng xe", OCR often concatenates (e.g. "XăngRON95"), so use substring match
            if category == "Xăng xe":
                if kw.lower() in text_lower:
                    score += 1
            else:
                pattern = r'\b' + re.escape(kw.lower()) + r'\b'
                if re.search(pattern, text_lower):
                    score += 1
        
        if score > 0:
            scores[category] = score
    
    if scores:
        return max(scores, key=scores.get)
    return "Khác"


def format_price_value(value):
    
    if not value or not isinstance(value, str):
        return value
    
    # Remove any spaces
    value = value.strip()
    
    # Use robust parse_money logic to handle both 1.234,56 and 1,234.56
    # Instead of naive split, we strip non-numeric and parse
    try:
        # Check if it has decimal part like ,00 or .00 at end
        s = value.strip()
        import re
        if '.' in s and ',' in s:
            if s.rfind(',') > s.rfind('.'): # 1.234,56
                s = s[:s.rfind(',')]
            else: # 1,234.56
                s = s[:s.rfind('.')]
        elif re.search(r'[,.]\d{2}$', s) and not re.search(r'[,.]\d{3}$', s):
             # 2 digit decimal suffix
             s = s[:-3]
             
        # Now remove any remaining separators
        clean = s.replace('.', '').replace(',', '')
        num = int(clean)
        return f"{num:,}" # Format with comma
    except:
        return value

COMMON_UNITS = {
    "CÁI", "CHIẾC", "BỘ", "GÓI", "HỘP", "THÙNG", "BAO", "CHAI", "LON", "LÍT", "LIT", "KG", "GRAM", "GM", "MÉT", 
    "M", "M2", "M3", "CUỘN", "TẤM", "THANH", "VIÊN", "VỈ", "TỜ", "QUYỂN", "CUỐN", "RAM", "CẶP", "ĐÔI", 
    "DĨA", "ĐĨA", "PHẦN", "THỐ", "TÔ", "CHÉN", "LY", "CỐC", "SUẤT", "KIM", "CHẬU", "CÂY", "GIỜ", "NGÀY", "THÁNG", 
    "NĂM", "LẦN", "CHUYẾN", "LƯỢT", "PHÚT", "KW", "KWH", "SỐ", "MÓN", "KỆ", "BỊCH", "NỒI", "CON", "PCS", "NGƯỜI"
}

# Keywords indicating item lines to skip (headers, footers, summaries)
JUNK_TEXT_KEYWORDS = [
    'stt', 'tên hàng', 'đơn vị tính', 'số lượng', 'thành tiền', 'người mua', 
    'ký bởi', 'trang', 'thuế suất', 'cộng tiền', 'tổng cộng', 'bằng chữ',
    'tiền thuế', 'serial', 'ký hiệu', 'mẫu số', 'vnd', 'chuyển khoản',
    'vat invoice', 'đơn vị bán', 'mã tra cứu', 'vat) rate)', 'vat rate', 
    'gtgt', 'rate)', 'amount)', 'rate%)', 'tên h', 'đơ n v', 's ố l', 
    'vị tính', 'sau thuế', 'chiết khấu', 'a b c'
]

# Keywords for detecting surcharge/fee items
SURCHARGE_KEYWORDS = ['phụ thu', 'phí dịch vụ', 'phí phục vụ', 'service charge', 'surcharge']


def is_junk_text(text):
    """Check if text is a header/footer/summary line that should be skipped."""
    if not text or len(text) < 2:
        return True
    t = text.lower()
    if re.match(r'^([A-Z]\s+)+[A-Z]$', text):
        return True
    if re.match(r'^[\d\s()=x+]+$', text):
        return True
    if any(w in t for w in JUNK_TEXT_KEYWORDS):
        return True
    if len(t) > 50 and sum(1 for c in t if c in '()') > 4:
        return True
    return False


def parse_money(value):
    """Parse Vietnamese money string to integer. Returns None if invalid."""
    if not value:
        return None
    try:
        s = str(value).strip()
        
        # Robust Logic: Check order of separators if both exist
        if '.' in s and ',' in s:
            last_dot = s.rfind('.')
            last_comma = s.rfind(',')
            if last_comma > last_dot:  # 1.234,56 -> , is decimal
                s = s[:last_comma]
            else:  # 1,234.56 -> . is decimal
                s = s[:last_dot]
        else:
            # Only one separator type - check for 2-digit decimal suffix
            import re
            if re.search(r'[,.]\d{2}$', s) and not re.search(r'[,.]\d{3}$', s):
                s = s[:-3]
        
        clean = s.replace('.', '').replace(',', '')
        return int(clean)
    except (ValueError, TypeError):
        return None


def format_money(value):
    """Format integer as money string with comma separator."""
    if value is None:
        return ""
    return f"{value:,}"


def parse_vietnamese_number(value):
    """Parse Vietnamese number format (dot as thousand separator, comma as decimal)."""
    if not value:
        return 0
    try:
        return float(str(value).replace('.', '').replace(',', '.'))
    except (ValueError, TypeError):
        return 0


def ocr_pdf_to_text(pdf_source, filename=None):
    """
    Use OCR to extract text from scanned PDF.
    Returns extracted text or empty string if OCR fails.
    """
    if not OCR_AVAILABLE:
        print("  OCR not available (pytesseract/pdf2image not installed)")
        return ""
    
    try:
        import tempfile
        
        # Convert PDF to images
        if isinstance(pdf_source, str):
            # File path - use directly
            if POPPLER_PATH:
                images = convert_from_path(pdf_source, dpi=300, poppler_path=POPPLER_PATH)
            else:
                images = convert_from_path(pdf_source, dpi=300)
        else:
            # BytesIO stream - save to temp file first
            # IMPORTANT: Seek to beginning before reading
            pdf_source.seek(0)
            pdf_bytes = pdf_source.read()
            pdf_source.seek(0)  # Reset for potential future use
            
            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
                tmp.write(pdf_bytes)
                tmp_path = tmp.name
            
            try:
                if POPPLER_PATH:
                    images = convert_from_path(tmp_path, dpi=300, poppler_path=POPPLER_PATH)
                else:
                    images = convert_from_path(tmp_path, dpi=300)
            finally:
                os.unlink(tmp_path)  # Clean up temp file
        
        # OCR each page
        full_text = ""
        for i, image in enumerate(images):
            # Try Vietnamese+English first, fallback to English only if vie not available
            try:
                text = pytesseract.image_to_string(image, lang='vie+eng')
            except Exception:
                try:
                    text = pytesseract.image_to_string(image, lang='vie')
                except Exception:
                    text = pytesseract.image_to_string(image, lang='eng')
            full_text += text + "\n"
        
        return full_text
    except Exception as e:
        print(f"  OCR error: {e}")
        return ""


def _parse_vietnamese_words_to_number(text):
    """Parse Vietnamese word-form amount to integer. E.g. 'bảy trăm nghìn' -> 700000"""
    text = text.strip().lower()
    text = re.sub(r'\b(đồng|dong|chẵn|chan|chấn|lẻ|le|và|va)\b', '', text).strip()
    
    digit_map = {
        'không': 0, 'một': 1, 'mot': 1, 'hai': 2, 'ba': 3, 'bốn': 4, 'bon': 4,
        'năm': 5, 'nam': 5, 'lăm': 5, 'sáu': 6, 'sau': 6, 'bảy': 7, 'bay': 7,
        'tám': 8, 'tam': 8, 'chín': 9, 'chin': 9, 'linh': 0, 'lẻ': 0, 'le': 0,
    }
    
    tokens = text.split()
    if not tokens:
        return 0
    
    # Parse group by group. Each group can have: [digit] trăm [digit] mươi [digit]
    # Groups are separated by nghìn/triệu/tỷ
    result = 0
    hundreds = 0   # accumulated hundreds value
    pending = 0    # last digit waiting for mươi/trăm
    
    def flush_group(multiplier=1):
        nonlocal result, hundreds, pending
        group_val = hundreds + pending
        result += group_val * multiplier
        hundreds = 0
        pending = 0
    
    for token in tokens:
        if token in digit_map:
            pending = digit_map[token]
        elif token in ('mười', 'muoi', 'mươi'):
            if pending == 0:
                hundreds += 10
            else:
                hundreds += pending * 10
                pending = 0
        elif token in ('trăm', 'tram'):
            hundreds += pending * 100
            pending = 0
        elif token in ('nghìn', 'nghin', 'ngàn', 'ngan'):
            if hundreds == 0 and pending == 0:
                pending = 1
            flush_group(1000)
        elif token in ('triệu', 'trieu'):
            if hundreds == 0 and pending == 0:
                pending = 1
            flush_group(1000000)
        elif token in ('tỷ', 'ty', 'tỉ'):
            if hundreds == 0 and pending == 0:
                pending = 1
            flush_group(1000000000)
    
    # Flush remaining
    flush_group(1)
    return result


def extract_ocr_invoice_fields(text, filename=None):
    """Extract invoice fields from OCR text (simpler patterns for OCR quality)."""
    data = {}
    
    # Debug: print relevant parts for money extraction
    text_lower = text.lower()
    print(f"  OCR TEXT (looking for money patterns):")
    # Find and print lines with money-related keywords
    for line in text.split('\n'):
        line_lower = line.lower().strip()
        if any(kw in line_lower for kw in ['tiền', 'tien', 'hang', 'hàng', 'thuế', 'thue', 'gtgt', 'cộng', 'cong', 'tổng', 'tong', 'thanh toán', 'thanh toan', 'cxc']):
            print(f"    >> {line.strip()}")
    
    # Ký hiệu
    serial_match = re.search(r'[Kk]ý\s*hiệu[:\s]*([A-Z0-9]+)', text)
    if serial_match:
        data["Ký hiệu"] = serial_match.group(1)
        
    # Đơn vị bán (Seller) - New for OCR
    seller_candidates = []
    lines = text.split('\n')
    for i, line in enumerate(lines[:15]): # Check first 15 lines
        line_clean = line.strip()
        # Common prefix for companies
        if re.match(r'^(CÔNG TY|CHI NHÁNH|DNTN|TRUNG TÂM|HỘ KINH DOANH|CỬA HÀNG)', line_clean, re.IGNORECASE):
            seller_candidates.append(line_clean)
        elif 'PETROLIMEX' in line_clean.upper():
            seller_candidates.append(line_clean)
    
    if seller_candidates:
        # Pick the longest one logic or first one
        raw_seller = seller_candidates[0]
        # Cleanup "Ký hiệu: ..." from snippet if attached
        if "Ký hiệu:" in raw_seller:
            raw_seller = raw_seller.split("Ký hiệu:")[0].strip()
        data["Đơn vị bán"] = raw_seller
    else:
        # Fallback: if 'bán' or 'seller' keyword exists
        pass
    
    # Số hóa đơn - multiple patterns
    inv_patterns = [
        r'[Ss][oố]\s*hóa\s*đơn[:\s]+(\d{5,})',
        r'[Ss]ố\s*(?:HĐ)[:\s]*(\d+)',
        r'[Ss][oố][:\s]+(\d{6,})',
        r'[Nn]o\.?[:\s]*(\d{5,})',
        r'Invoice No\.?[:\s]*(\d+)',
        r'[Ss][ốo]\s*[(/]?\s*No\.?\s*[)/]?[:\s]*(\d+)', # Matches "Số (No.): ..."
        r'\(\s*VAT\s*INVOICE\s*\)[:\s]*(\d+)',
        r'Số:\s*(\d+)',
    ]
    for p in inv_patterns:
        m = re.search(p, text)
        if m:
            num = m.group(1)
            if not re.match(r'^(18|19|09|08|07|06|05|03|02|01)\d{6,}', num):
                data["Số hóa đơn"] = num
                break
    
    # Fallback: from filename
    if "Số hóa đơn" not in data and filename:
        fn_match = re.search(r'_(\d{5,})', filename)
        if fn_match:
            data["Số hóa đơn"] = fn_match.group(1)
    
    # Ngày hóa đơn
    date_match = re.search(r'[Nn]gày\s*(\d{1,2})\s*tháng\s*(\d{1,2})\s*năm\s*(\d{4})', text)
    if date_match:
        data["Ngày hóa đơn"] = f"{date_match.group(1)}/{date_match.group(2)}/{date_match.group(3)}"
    else:
        # Xang3 fix: un-accented OCR date "Ngay 15 thang 02 nam 2026"
        date_match_noaccent = re.search(r'[Nn]ga[yỳ]\s*(\d{1,2})\s*thang\s*(\d{1,2})\s*nam\s*(\d{4})', text)
        if date_match_noaccent:
            data["Ngày hóa đơn"] = f"{date_match_noaccent.group(1)}/{date_match_noaccent.group(2)}/{date_match_noaccent.group(3)}"
        else:
            date_match2 = re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})', text)
            if date_match2:
                data["Ngày hóa đơn"] = f"{date_match2.group(1)}/{date_match2.group(2)}/{date_match2.group(3)}"
    
    # MST bên bán
    ignore_mst = ['0106869738', '0100684378', '0101245171', '0305482862', '0103243195', '0101360697']
    mst_patterns = [
        r'[Mm]a\s*số\s*thuế[:\s]*([\d\-\u00AD\s]+)',
        r'MST[:\s]*([\d\-\u00AD\s]+)',
        r'[Mm]ã\s*số\s*thuế[:\s]*([\d\-\u00AD\s]+)',
        r'[Mm]a\s*s[eoc]\s*thu[eé][:\s]*([\d\-\u00AD\s]+)', # Handle "Ma sé thué"
        r'[Mm]a\s*s.\s*thu.[:\s]*([\d\-\u00AD\s]+)',
        r'tax\s*code[:\s]*([\d\-\u00AD\s]+)',
    ]
    
    found_mst = None
    for p in mst_patterns:
        if found_mst: break
        for m in re.finditer(p, text):
            raw_val = m.group(1)
            val = raw_val.replace(' ', '').replace('.', '').replace('-', '').replace('\u00AD', '').strip()
            # Valid length
            if len(val) < 10 or len(val) > 14: continue
            
            # Check blacklist
            if any(ign in val for ign in ignore_mst):
                print(f"  [OCR-MST] Ignored blacklisted: {val}")
                continue
                
            # Check context (line content)
            start = text.rfind('\n', 0, m.start()) + 1
            end = text.find('\n', m.end())
            if end == -1: end = len(text)
            line_content = text[start:end].lower()
            
            # Ignore context keywords (Provider or Buyer)
            if any(kw in line_content for kw in ['giải pháp', 'phần mềm', 'cung cấp bởi', 'phát hành bởi', 'created by', 'signature', 'ký bởi', 'bkav', 'ehoadon', 'mua hàng', 'người mua', 'đơn vị mua']):
                 print(f"  [OCR-MST] Ignored bad context: {val} in '{line_content.strip()[:50]}...'")
                 continue
            
            # Found a good one!
            found_mst = val
            data["Mã số thuế"] = found_mst
            print(f"  [OCR-MST] Accepted: {found_mst}")
            break
            
    # Fallback: Petrolimex specific hardcoded patterns caused by OCR errors?
    # Based on log: "Ma sé thué: 0300555450" but regex above might miss due to spacing oddities.
    if not data.get("Mã số thuế"):
        # Just look for the specific sequence "Ma se thue" in normalized text
        norm = text.lower().replace('é', 'e').replace('ú', 'u').replace('ế', 'e')
        # Matches: "ma se thue: 0300555450"
        m_ocr = re.search(r'(?:ma se thue|ma so thue|mst)[^0-9]*([0-9]{10,14})', norm)
        if m_ocr:
             cand = m_ocr.group(1)
             if not any(ign in cand for ign in ignore_mst):
                 data["Mã số thuế"] = cand
                 print(f"  [OCR-MST] Accepted via fallback: {cand}")
    
    # Số tiền trước thuế (Petrolimex specific patterns from cloud OCR)
    before_patterns = [
        r'ông\s*tiên\s*hàng[:\s]*([\d\.,]+)',        # "ông tiên hàng: 462.963"
        r'ông\s*tiên\s*hang[:\s]*([\d\.,]+)',        # "ông tiên hang: 481.787"
        r'[Cc]ộng\s*tiền\s*hàng[:\s]*([\d\.,]+)',
        r'[Tt]iền\s*hàng[:\s]*([\d\.,]+)',
    ]
    for p in before_patterns:
        m = re.search(p, text)
        if m:
            val = m.group(1).strip()
            # Skip if looks like year (2025, 2026) or too short
            if not re.match(r'^20[0-9]{2}$', val) and len(val) >= 3:
                data["Số tiền trước Thuế"] = val
                break
    
    # VAT (Petrolimex specific - from cloud OCR)
    # Pattern: "lên thuê GTGT (8% ) 38.543" or "lên thuê GTGT ( 8% )"
    vat_patterns = [
        r'lên\s*thuê\s*GTGT\s*\(\s*\d+\s*%?\s*\)\s*([\d\.,]+)',   # "lên thuê GTGT (8% ) 38.543"
        r'ién\s*thuê\s*GTGT\s*\(\s*\d+\s*%?\s*\)\s*([\d\.,]+)',   # OCR variant
        r'[Tt]iền\s*thuế\s*GTGT[:\s]*([\d\.,]+)',
        r'thuê\s*GTGT\s*\(\s*8\s*%?\s*\)\s*([\d\.,]+)',
        r'GTGT\s*\(\s*\d+\s*%?\s*\)\s*([\d\.,]+)',
        r'[Cc]XC[:\s]*([\d\.,]+)',
        r'thuế\s*GTGT[:\s]*([\d\.,]+)',
    ]
    for p in vat_patterns:
        m = re.search(p, text)
        if m:
            data["Tiền thuế"] = m.group(1)
            break
    
    # Tax rate detection (Petrolimex uses 8%)
    # For gas stations, default to 8% VAT
    if 'petrolimex' in text.lower() or 'xăng' in text.lower() or 'ron 95' in text.lower():
        if data.get("Tiền thuế"):
            data["Thuế 8%"] = data["Tiền thuế"]
    elif re.search(r'8\s*%', text):
        data["Thuế 8%"] = data.get("Tiền thuế", "")
    elif re.search(r'10\s*%', text):
        data["Thuế 10%"] = data.get("Tiền thuế", "")
    
    # Tổng tiền sau thuế (Petrolimex specific - from cloud OCR)
    # Pattern: "ông sô tiên thanh toán: 800.083"
    total_patterns = [
        r'ông\s*sô\s*tiên\s*thanh\s*toán[:\s]*([\d\.,]+)',        # "ông sô tiên thanh toán: 800.083"
        r'[Tt]ổng\s*(?:số\s*)?(?:cộng|tiền)\s*thanh\s*toán[:\s]*([\d\.,]+)',
        r'thanh\s*toán[:\s]*([\d\.,]+)',
        r'[Tt]ổng\s*(?:cộng|tiền)[:\s]*([\d\.,]+)',
    ]
    for p in total_patterns:
        m = re.search(p, text)
        if m:
            val = m.group(1).strip()
            if not re.match(r'^20[0-9]{2}$', val) and len(val) >= 3:
                data["Số tiền sau"] = val
                break
    
    # Xang2 fix: Fallback - parse Vietnamese word-based amount when OCR misses numeric value
    # "Tổng số tiền thanh toán bằng chữ: Bảy trăm nghìn đồng" = 700,000
    if not data.get("Số tiền sau"):
        chu_match = re.search(r'(?:bằng chữ|bang chu)[:\s]*(.+?)(?:đồng|dong|$)', text, re.IGNORECASE)
        if chu_match:
            words = chu_match.group(1).strip().lower()
            vn_num = _parse_vietnamese_words_to_number(words)
            if vn_num and vn_num > 0:
                data["Số tiền sau"] = f"{vn_num:,.0f}".replace(',', '.')
                print(f"  [OCR-WORDS] Parsed '{words}' = {vn_num}")
    
    # Auto-calculate missing values (for Petrolimex 8% VAT)
    def parse_money(s):
        """Convert string like '481.787' or '1.820.000,00' to int, handling decimals"""
        if not s:
            return 0
        s = str(s).strip()
        # Robust Logic: Check order of separators if both exist
        if '.' in s and ',' in s:
            last_dot = s.rfind('.')
            last_comma = s.rfind(',')
            if last_comma > last_dot: # 1.234,56 -> , is decimal
                s = s[:last_comma]
            else: # 1,234.56 -> . is decimal
                s = s[:last_dot]
        else:
            # Only one separator type
            if re.search(r'[,.]\d{2}$', s) and not re.search(r'[,.]\d{3}$', s):
                 s = s[:-3]
        s = s.replace(',', '').replace('.', '')
        try:
            return int(s)
        except:
            return 0
    
    def format_money(n):
        """Format number back to string with dots as thousands separator"""
        return f"{n:,.0f}".replace(',', '.')
    
    before_tax = parse_money(data.get("Số tiền trước Thuế", ""))
    vat = parse_money(data.get("Tiền thuế", ""))
    total = parse_money(data.get("Số tiền sau", ""))
    
    # If missing total but have before_tax and vat, calculate
    if not total and before_tax and vat:
        total = before_tax + vat
        data["Số tiền sau"] = format_money(total)
        print(f"  [AUTO-CALC] Total = {before_tax} + {vat} = {total}")
    
    # If missing total but have before_tax (assume 8% VAT for Petrolimex)
    elif not total and before_tax and 'petrolimex' in text.lower():
        vat = int(round(before_tax * 0.08))
        total = before_tax + vat
        data["Số tiền sau"] = format_money(total)
        data["Tiền thuế"] = format_money(vat)
        data["Thuế 8%"] = format_money(vat)
        print(f"  [AUTO-CALC] VAT 8% = {vat}, Total = {total}")
    
    # If missing before_tax but have total (assume 8% VAT for Petrolimex)
    elif not before_tax and total and 'petrolimex' in text.lower():
        before_tax = int(round(total / 1.08))
        vat = total - before_tax
        data["Số tiền trước Thuế"] = format_money(before_tax)
        if not data.get("Tiền thuế"):
            data["Tiền thuế"] = format_money(vat)
            data["Thuế 8%"] = format_money(vat)
        if not data.get("Tiền thuế"):
            data["Tiền thuế"] = format_money(vat)
            data["Thuế 8%"] = format_money(vat)
        print(f"  [AUTO-CALC] Before tax = {before_tax}, VAT = {vat}")
        
    # If missing VAT but have before_tax and total
    elif not vat and before_tax and total:
        vat = total - before_tax
        data["Tiền thuế"] = format_money(vat)
        # Infer rate
        rate = round(vat / before_tax, 2)
        if rate == 0.08:
            data["Thuế 8%"] = format_money(vat)
        elif rate == 0.10:
            data["Thuế 10%"] = format_money(vat)
        elif rate == 0.05:
            data["Thuế 5%"] = format_money(vat)
        elif 'petrolimex' in text.lower(): # Standardize Petrolimex to 8% if ambiguous
             data["Thuế 8%"] = format_money(vat)
        else:
             data["Thuế khác"] = format_money(vat)
        print(f"  [AUTO-CALC] Inferred VAT = {vat} (Rate ~{rate})")
    
    # Mã tra cứu
    lookup_match = re.search(r'(?:Mã|Ma)\s*(?:tra|tro)\s*(?:cứu|cuiu|cuu)[:\s]*([A-Z0-9*]+)', text, re.IGNORECASE)
    if not lookup_match:
         # Try finding code at bottom
         lookup_match = re.search(r'\b([A-Z0-9]{10,})\b', text) # Simple long code
    if lookup_match and len(lookup_match.group(1)) > 5:
        data["Mã tra cứu"] = lookup_match.group(1)
        
    # Mã CQT (New)
    cqt_match = re.search(r'(?:Mã|Ma)\s*(?:CQT|cơ\s*quan\s*thuế)[:\s]*([A-Z0-9\-]+)', text, re.IGNORECASE)
    if cqt_match:
        data["Mã CQT"] = cqt_match.group(1)
    
    # Link
    link_match = re.search(r'(https?://[^\s]+)', text)
    if link_match:
        data["Link lấy hóa đơn"] = link_match.group(1)
    
    return data


def extract_services_from_text(full_text):
    """Extract service/product details with qty, unit_price, and amount."""
    services = []
    lines = full_text.split('\n')

    for line_idx, line in enumerate(lines):
        line = line.strip()
        if not line:
            continue
            
        first_token = line.split()[0] if line.split() else ""
        if not first_token.isdigit():
            continue
            
        # Ignore column headers "1 2 3 4 5"
        if re.match(r'^[\d\s.,|()x=+]+$', line): continue
        if re.match(r'^[A-Z\s]+[\d\s=x]+$', line): continue
             
        # Must start with number (1-3 digits) - STT
        match_start = re.match(r'^(\d{1,3})[._\-\s|]+', line)
        if not match_start: continue
        
        # Find all number blocks
        num_pattern = r'(\d+(?:[.,]\d+)*)'
        all_nums = list(re.finditer(num_pattern, line))
        
        if len(all_nums) < 3:  # Need at least STT + qty + amount
            # Exception: surcharge items may have only STT + amount (2 numbers)
            line_lower = line.lower()
            is_surcharge_line = any(kw in line_lower for kw in SURCHARGE_KEYWORDS)
            if not (is_surcharge_line and len(all_nums) == 2):
                continue
        
        # Detect Tax Rate explicitly (e.g. "8%", "10%")
        tax_rate = None
        rate_match = re.search(r'\b(0|5|8|10)\s*%', line)
        if rate_match:
            tax_rate = rate_match.group(1)
        
        stt_end = len(match_start.group(0))
        
        # Strategy: Find the LAST UNIT word (from COMMON_UNITS) that is followed by numbers
        # This handles cases where unit words appear in description (like "từ ngày 15/12")
        
        tokens = line.split()
        
        line_after_stt = line[stt_end:]
        unit_idx = -1  # index of unit word in tokens
        
        # Find LAST unit word that has numbers after it
        for i in range(len(tokens) - 1, -1, -1):  # iterate backwards
            tok = tokens[i]
            # Ignore THANH as unit (usually start of name like Thanh long)
            if tok.upper() in COMMON_UNITS and tok.upper() != 'THANH':
                # Check if there are numbers after this position
                remaining = ' '.join(tokens[i+1:])
                if re.search(r'\d+(?:[.,]\d+)*', remaining):
                    unit_idx = i
                    break
        
        if unit_idx == -1:
            # No unit found - find name boundary by looking for first "price-like" number
            # A price typically has 4+ digits or contains decimal separator
            nums_after_stt = [m for m in all_nums if m.start() > stt_end]
            
            # Find first number that looks like a price (>= 1000 or has comma/dot)
            name_end_pos = len(line)
            for m in nums_after_stt:
                num_str = m.group(1)
                # Check if it's a price number: contains . or , separator OR is >= 4 digits
                if '.' in num_str or ',' in num_str or len(num_str.replace('.', '').replace(',', '')) >= 4:
                    name_end_pos = m.start()
                    break
            
            if name_end_pos <= stt_end:
                continue
            name_part = line[stt_end:name_end_pos].strip()
            nums = [m.group(1) for m in nums_after_stt]
        else:
            # Unit found - take name as tokens before unit, numbers after unit
            name_tokens = tokens[:unit_idx]
            
            # Special case: if the last token before unit is a 4-digit year (e.g., "2025"), include it in name
            # This handles cases like "tháng 11 năm 2025 Tháng 1 1.260.000"
            if unit_idx > 0 and re.match(r'^\d{4}$', tokens[unit_idx - 1]):
                # Keep the year in name, effectively making unit_idx point to one position after
                pass  # name_tokens already includes the year since it's tokens[:unit_idx]
            
            # Instead of offset calculation which is error-prone, just take tokens after unit
            remaining_tokens = tokens[unit_idx+1:]
            remaining_text = " ".join(remaining_tokens)
            
            nums_iter = re.finditer(r'(\d+(?:[.,]\d+)*)', remaining_text)
            nums = [m.group(1) for m in nums_iter]
            
            if len(nums) < 2:
                # Check if it's a surcharge item before skipping
                temp_name = ' '.join(name_tokens).lower()
                is_surcharge = any(kw in temp_name for kw in SURCHARGE_KEYWORDS)
                if not (is_surcharge and len(nums) == 1):
                    continue
                    
            name_part = ' '.join(name_tokens)
        
        # Special handling for surcharge/fee items with only one number (just amount, no qty/price)
        # e.g., "19 Phụ thu 171.500"
        is_surcharge_item = any(kw in name_part.lower() for kw in SURCHARGE_KEYWORDS)
        
        if len(nums) < 2:
            if is_surcharge_item and len(nums) == 1:
                # Surcharge with single amount - use it as both price and amount, qty = 1
                nums = ['1', nums[0], nums[0]]  # [qty, unit_price, amount]
            else:
                continue
        
        # Clean name - remove unit words from start and end
        tokens = name_part.split()
        # Remove from end - only exact match units
        while tokens and tokens[-1].upper() in COMMON_UNITS:
            tokens = tokens[:-1]
        # Remove from start (rare but can happen with multi-line merge) - only exact match
        # EXCEPTION: "THANH" is a unit but also start of "Thanh long", "Thanh toán"... so don't remove it at start
        while tokens and tokens[0].upper() in COMMON_UNITS and tokens[0].upper() != "THANH":
            tokens = tokens[1:]
        # Also clean partial unit patterns like "Phần）" or "）Nấm"  
        if tokens:
            # Clean first token if it's just punctuation + text
            first = tokens[0]
            if first.startswith('）') or first.startswith(')'):
                tokens[0] = first[1:].strip()
            # Clean last token if it ends with LONG unit (3+ chars) to avoid cutting "Nấm" -> "Nấ"
            last = tokens[-1] if tokens else ''
            for unit in COMMON_UNITS:
                if len(unit) >= 3 and last.upper().endswith(unit) and len(last) > len(unit):
                    tokens[-1] = last[:-len(unit)].rstrip('（(')
                    break
            # Clean tokens that are JUST unit+bracket like "Phần）" or "Phần)"
            # Only remove if it's strictly unit+bracket, to avoid removing "Nửa phần）"
            # Clean tokens that are JUST unit+bracket like "Phần）" or "Phần)" removal logic removed 
            # as it truncated valid name parts like "Bắp Mỹ（Nửa phần）" where "phần）" is part of the name.
            pass
        name_part = " ".join(tokens)
        
        # Remove STT from the start of name_part if detected
        # This prevents cases like "4 Phần" becoming "Đậu phụ... 4 Phần" after merge
        stt_val = match_start.group(1)
        if name_part.startswith(stt_val):
             name_part = name_part[len(stt_val):].strip()
        
        # Multi-line description handling: merge PREVIOUS and NEXT lines if they look like continuations
        
        # Check PREVIOUS lines for description prefix - but only if current name looks incomplete
        # Conditions for needing prev line: starts lowercase, starts with '(' or ')', or is very short
        # Ignore leading STT number for this check
        check_name = re.sub(r'^\d+\s+', '', name_part).strip()
        first_char = check_name[0] if check_name else ''
        needs_prev = (
            len(check_name) < 5 or  # Very short name
            (first_char and first_char.islower()) or  # Starts lowercase = continuation
            (first_char == '(' or first_char == ')') or  # Starts with paren
            (')' in check_name[:10])  # Has closing paren early
        )
        
        if needs_prev:
            prev_parts = []
            for offset in range(1, 3):  # Check up to 2 lines back
                if line_idx - offset >= 0:
                    prev_line = lines[line_idx - offset].strip()
                    
                    # Stop if: empty, STT line, junk, or has too many numbers
                    if not prev_line or len(prev_line) < 2:
                        break
                    match_stt = re.match(r'^(\d{1,3})[._\-\s|]+', prev_line)
                    if match_stt:
                         # If STT line contains ONLY STT and text (no other numbers), merge it
                         stt_len = len(match_stt.group(0))
                         rest_of_line = prev_line[stt_len:].strip()
                         
                         # Check if remaining part has numbers (likely prices) -> then it's a separate item -> Break
                         if re.search(r'\d', rest_of_line):
                             break
                             
                         # If valid text, use it (strip STT)
                         prev_line = rest_of_line
                    if is_junk_text(prev_line):
                        break
                        
                    # Check nums, but allow dates (DD/MM/YYYY or DD-MM-YYYY)
                    # Remove dates from line before counting nums
                    temp_line = re.sub(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', '', prev_line)
                    
                    # Only count "price-like" numbers (with comma/period separators) as stop conditions
                    # This allows alphanumeric codes like "63A 17235" (vehicle plates) to pass through
                    price_nums = re.findall(r'\d{1,3}(?:[.,]\d{3})+(?:[.,]\d{2})?', temp_line)
                    simple_nums = re.findall(r'\b\d{4,}\b', temp_line)  # Long standalone numbers (4+ digits)
                    
                    if price_nums or len(simple_nums) > 1:
                        break
                    if any(x in prev_line.lower() for x in ['cộng tiền', 'tổng cộng', 'thuế', 'thành tiền']):
                        break
                    
                    # STOP if prev_line looks like the TAIL of a previous item
                    if (prev_line.endswith(')') or prev_line.endswith('）')) and re.search(r'[a-zA-Z]', prev_line):
                        if '(' not in prev_line and '（' not in prev_line: 
                            break
                    
                    # STOP merging if we hit an English paren line that precedes a Vietnamese line
                    # Case Item 7: "Thêm cơm trắng" (collected) <-- "(Sichuan tofu)" (checking). Stop.
                    # Case Item 5: "(Braised..." (collected) <-- "Chân nấm..." (checking). Continue.
                    if prev_parts:
                        last_collected = prev_parts[0] # This is the line conceptually "below" the current prev_line
                        is_vietnamese_line = last_collected[0].isupper() and not last_collected.startswith('(')
                        is_english_paren = prev_line.startswith('(') and re.search(r'[a-zA-Z]', prev_line)
                        if is_vietnamese_line and is_english_paren:
                            break

                    prev_parts.insert(0, prev_line)  # Insert at beginning to maintain order
                else:
                    break
            
            if prev_parts:
                name_part = " ".join(prev_parts) + " " + name_part
        
        # Check NEXT lines for description suffix
        last_char = name_part[-1] if name_part else ''
        
        # Detect if next line is a parenthetical suffix
        next_line_is_suffix = False
        if line_idx >= 0 and line_idx + 1 < len(lines):
            peek_line = lines[line_idx + 1].strip()
            if peek_line.startswith('(') and re.search(r'[a-zA-Z]', peek_line):
                next_line_is_suffix = True
            elif peek_line.startswith('(') and re.search(r'(ngày|từ|đến|tháng|năm)', peek_line, re.IGNORECASE):
                next_line_is_suffix = True

        has_unclosed_paren = (name_part.count('(') > name_part.count(')')) or (name_part.count('（') > name_part.count('）'))
        
        needs_next = (
            last_char in '(-（' or
            name_part.rstrip().endswith('(') or
            name_part.rstrip().endswith('（') or
            has_unclosed_paren or
            next_line_is_suffix
        )
        
        if needs_next:
            next_parts = []
            for offset in range(1, 4):  # Check up to 3 lines ahead
                if line_idx >= 0 and line_idx + offset < len(lines):
                    next_line = lines[line_idx + offset].strip()
                    
                    # Stop if: empty, STT line, junk, or has too many numbers
                    if not next_line or len(next_line) < 2:
                        break
                    if re.match(r'^\d{1,3}[._\-\s|]+', next_line):
                        break
                    if is_junk_text(next_line):
                        break
                    # Remove date patterns before counting numbers
                    temp_next = re.sub(r'\d{1,2}\s*[/-]\s*\d{1,2}\s*[/-]\s*\d{2,4}', '', next_line)
                    nums_in_next = re.findall(r'\d+(?:[.,]\d+)*', temp_next)
                    if len(nums_in_next) > 1:
                        break
                    if any(x in next_line.lower() for x in ['cộng tiền', 'tổng cộng', 'thuế', 'thành tiền']):
                        break
                    
                    # STOP if next_line starts a new item (Vietnamese text, not closing paren)
                    if next_line[0].isupper() and next_line[0] != '(':
                        matches_closing_paren = has_unclosed_paren and ('）' in next_line or ')' in next_line)
                        if not matches_closing_paren:
                            if re.search(r'[àáạảãâầấậẩẫăằắặẳẵèéẹẻẽêềếệểễìíịỉĩòóọỏõôồốộổỗơờớợởỡùúụủũưừứựửữỳýỵỷỹđ]', next_line, re.IGNORECASE):
                                break
                    
                    next_parts.append(next_line)
                else:
                    break
            
            if next_parts:
                name_part = name_part + " " + " ".join(next_parts)
        
        name_part = name_part.strip()
        
        # Clean up junk prefixes and suffixes
        prefixes_to_remove = ["GTGT", "VAT) rate)", "VAT rate", "Rate)", "A B C", "khấu", "KHẤU", 
                             "Phần）", "Phần)", "PHẦN）", "PHẦN)", "ĐVT:", "ĐVT"]
        for prefix in prefixes_to_remove:
            if name_part.startswith(prefix + " "):
                name_part = name_part[len(prefix):].strip()
            if name_part.startswith(prefix):
                name_part = name_part[len(prefix):].strip()
        name_part = re.sub(r'^(?:[A-C]\s)+[\d\s=x]+', '', name_part)
        name_part = re.sub(r'^[\d\s=x+]+(\s|$)', '', name_part)
        
        # Remove trailing junk like "- 6 + 7 9 10 = 8 x 9" or "Nồi 2 400.000"
        # Pattern: ends with numbers/operators/units that look like column data
        name_part = re.sub(r'\s+[\-\+]\s*[\d\s=x\+\-]+$', '', name_part)
        # Remove trailing numbers ONLY if they're not:
        # 1. A 4-digit year (like 2025)  
        # 2. Preceded by price/value keywords (like "mệnh giá 20.000")
        trailing_match = re.search(r'(\S+)\s+(\d+[\s.,\d]*)$', name_part)
        if trailing_match:
            word_before = trailing_match.group(1).lower()
            num_part = trailing_match.group(2).strip().split()[0] if trailing_match.group(2).strip() else ""
            keep_number = False
            # Keep 4-digit years
            if len(num_part) == 4 and num_part.isdigit():
                keep_number = True
            # Keep if preceded by price/value keywords
            price_keywords = ['giá', 'gia', 'mệnh', 'số', 'phòng', 'room', 'no', 'no.']
            if word_before in price_keywords:
                keep_number = True
            if not keep_number:
                name_part = name_part[:trailing_match.start(2)].strip()
        
        name_part = name_part.strip()
        
        # FINAL cleanup: remove unit words that appeared after merge
        final_tokens = name_part.split()
        while final_tokens and final_tokens[-1].upper() in COMMON_UNITS:
            final_tokens = final_tokens[:-1]
        # EXCEPTION: "THANH" at start
        while final_tokens and final_tokens[0].upper() in COMMON_UNITS and final_tokens[0].upper() != "THANH":
            final_tokens = final_tokens[1:]
        name_part = " ".join(final_tokens)
        
        # Skip if name is junk
        if len(name_part) < 3:
            continue
        if is_junk_text(name_part):
            continue
        if re.match(r'^[\d\s]+$', name_part):
            continue
        if re.match(r'^[\d\s=x\+\-\.,()\[\]]+$', name_part):
            continue
        
        # Determine qty, unit_price, amount
        # Handle different formats:
        # - Standard (4-5 nums): Qty, Price, Amount, [Rate%, VAT] - Thành tiền at position 3 (index 2)
        # - With discount (6+ nums): Qty, Price, Discount(0), Amount, Rate%, VAT, Total - if nums[2]=0, use nums[3]
        qty = ""
        unit_price = ""
        amount = ""
        
        if len(nums) >= 5:
            # Check if there's a discount column (nums[2] is 0 or 0,00)
            discount_val = nums[2].replace('.', '').replace(',', '').strip('0')
            has_discount_column = (discount_val == '' or discount_val == '0')
            
            qty = nums[0]
            unit_price = nums[1]
            
            # Smart Selection: Compare candidates nums[2] and nums[3] against Expected = Qty * Price
            # This handles cases like: Qty, Price, ServiceCharge, Amount (HĐ 63518)
            cand2 = nums[2]
            cand3 = nums[3] if len(nums) >= 4 else None
            
            try:
                q_val = float(qty.replace(',', '.').replace('.', '') if ',' in qty else qty)
                p_val = float(unit_price.replace('.', '').replace(',', '.'))
                if p_val < 100 and '.' in unit_price:
                    p_val = float(unit_price.replace('.', '').replace(',', ''))
                
                expected = q_val * p_val
                if expected == 0 and q_val == 0:
                    expected = p_val
                
                v2 = parse_vietnamese_number(cand2)
                v3 = parse_vietnamese_number(cand3) if cand3 else 0
                
                diff2 = abs(v2 - expected)
                diff3 = abs(v3 - expected) if cand3 else float('inf')
                
                # Select amount candidate closest to expected value
                if cand3 and diff3 < diff2 and v2 < (0.5 * expected):
                    amount = cand3
                elif has_discount_column and cand3:
                    amount = cand3
                else:
                    amount = cand2
            except:
                # Fallback to logic based on discount column
                if has_discount_column and len(nums) >= 4:
                    amount = nums[3]
                else:
                    amount = nums[2]
        elif len(nums) >= 3:
            qty = nums[0]
            unit_price = nums[1]
            amount = nums[2]
        elif len(nums) == 2:
            qty = nums[0]
            amount = nums[1]
            unit_price = amount
        else:
             continue
        
        # Heuristic fix for "Phí dịch vụ" case where Amount is misidentified as Tax Rate (e.g. 8)
        # If extracted Amount is very small (<= 100) and Unit Price is substantial (> 1000), swap/fix it.
        try:
            # Simple parse assuming checking for small integers vs large price
            a_str = amount.replace('.', '').replace(',', '')
            p_str = unit_price.replace('.', '').replace(',', '')
            if a_str.isdigit() and p_str.isdigit():
                 val_amount = float(a_str)
                 val_price = float(p_str)
                 # Check raw values (accounting for potential decimal scaling issues, but 8 vs 46800 is clear)
                 if val_amount <= 100 and val_price > 1000:
                      amount = unit_price
                      if qty == '0' or not qty:
                           qty = "1"
        except:
            pass
        
        services.append({
            "name": name_part,
            "qty": format_price_value(qty),
            "unit_price": format_price_value(unit_price),
            "amount": format_price_value(amount),
            "tax_rate": tax_rate
        })

    return services

def extract_invoice_data(pdf_source, filename=None):
    """
    Extract invoice data from a PDF file source.
    :param pdf_source: File path (str) or file-like object (BytesIO)
    :param filename: Original filename (if pdf_source is a stream)
    """
    
    def parse_money(s):
        """Convert string like '481.787' or '1.820.000,00' to int, handling decimals"""
        if not s:
            return 0
        s = str(s).strip()
        
        # Robust Logic: Check order of separators if both exist
        if '.' in s and ',' in s:
            last_dot = s.rfind('.')
            last_comma = s.rfind(',')
            if last_comma > last_dot: # 1.234,56 -> , is decimal
                s = s[:last_comma]
            else: # 1,234.56 -> . is decimal
                s = s[:last_dot]
        else:
            # Only one separator type
            # Case: 1.820.000 -> Integer
            # Case: 50,05 (decimal)
            # Logic: If ends with 2 digits decimal suffix -> remove
            if re.search(r'[,.]\d{2}$', s) and not re.search(r'[,.]\d{3}$', s):
                 s = s[:-3]
             
        s = s.replace(',', '').replace('.', '')
        try:
            return int(s)
        except:
            return 0
            
    def format_money(n):
        """Format number back to string with dots as thousands separator"""
        if n is None: return ""
        if isinstance(n, str) and not n.strip(): return ""
        try:
             # Use our robust parse_money to get float/int
             val = parse_money(n)
             return f"{val:,.0f}".replace(',', '.')
        except:
             return str(n)
             
    if isinstance(pdf_source, str):
        filename = os.path.basename(pdf_source)
    elif filename is None:
        filename = "Unknown.pdf"

    data = {
        "Tên file": filename,
        "Ngày hóa đơn": "",
        "Số hóa đơn": "",
        "Đơn vị bán": "",
        "Phân loại": "",
        "Số tiền trước Thuế": "",
        "Thuế 0%": "",
        "Thuế 5%": "",
        "Thuế 8%": "",
        "Thuế 10%": "",
        "Thuế khác": "",
        "Tiền thuế": "",
        "Số tiền sau": "",
        "Link lấy hóa đơn": "",
        "Mã tra cứu": "",
        "Mã số thuế": "",
        "Mã CQT": "",
        "Ký hiệu": "",
        "Phí PV": ""
    }
    # Store line items separately for multi-row expansion
    line_items = []
    
    try:
        # Read text directly from PDF using pdfplumber
        full_text = ""
        with pdfplumber.open(pdf_source) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    full_text += page_text + "\n"
        
        # Check if PDF is scanned (no text extracted)
        if not full_text.strip():
            print(f"  PDF has no text, trying OCR: {filename}")
            ocr_text = ocr_pdf_to_text(pdf_source, filename)
            if ocr_text.strip():
                # Use OCR extraction for scanned PDFs
                ocr_data = extract_ocr_invoice_fields(ocr_text, filename)
                for key, val in ocr_data.items():
                    if key in data and val:
                        data[key] = val
                
                # HD-OCR fix: If invoice number or MST is still empty, try high-DPI cropped OCR
                # Petrolimex invoices have data that's often unreadable at 300 DPI
                if (not data.get("Số hóa đơn") or not data.get("Mã số thuế")) and OCR_AVAILABLE:
                    try:
                        import tempfile
                        if isinstance(pdf_source, str):
                            hd_path = pdf_source
                        else:
                            pdf_source.seek(0)
                            pdf_bytes = pdf_source.read()
                            pdf_source.seek(0)
                            tmp = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
                            tmp.write(pdf_bytes)
                            tmp.close()
                            hd_path = tmp.name
                        
                        try:
                            if POPPLER_PATH:
                                hd_images = convert_from_path(hd_path, dpi=600, poppler_path=POPPLER_PATH)
                            else:
                                hd_images = convert_from_path(hd_path, dpi=600)
                            
                            if hd_images:
                                img = hd_images[0]
                                w, h = img.size
                                
                                # Invoice number: crop top-right 30% x 15%
                                if not data.get("Số hóa đơn"):
                                    crop_box = (int(w * 0.7), 0, w, int(h * 0.15))
                                    cropped = img.crop(crop_box)
                                    try:
                                        hd_text = pytesseract.image_to_string(cropped, lang='vie+eng', config='--psm 11')
                                    except Exception:
                                        hd_text = pytesseract.image_to_string(cropped, lang='eng', config='--psm 11')
                                    hd_match = re.search(r'(?:[Ss][ốoéô]|[Nn]o\.?)\s*[:\s]*(\d{4,})', hd_text)
                                    if hd_match:
                                        data["Số hóa đơn"] = hd_match.group(1)
                                        print(f"  [HD-OCR] Found invoice number: {hd_match.group(1)}")
                                
                                # MST: crop top 25% full width (seller MST is on the left)
                                if not data.get("Mã số thuế"):
                                    crop_top = img.crop((0, 0, w, int(h * 0.25)))
                                    try:
                                        hd_text_top = pytesseract.image_to_string(crop_top, lang='vie+eng', config='--psm 6')
                                    except Exception:
                                        hd_text_top = pytesseract.image_to_string(crop_top, lang='eng', config='--psm 6')
                                    ignore_mst = ['0106869738', '0100684378', '0101245171', '0305482862', '0103243195', '0101360697']
                                    # Find all MST matches, take first non-blacklisted
                                    for mst_m in re.finditer(r'(?:Mã\s*số\s*thuế|MST|Ma\s*s.\s*thu.)[:\s]*([\d\-\s]+)', hd_text_top, re.IGNORECASE):
                                        mst_val = mst_m.group(1).replace(' ', '').replace('-', '').strip()
                                        if len(mst_val) >= 10 and not any(ign in mst_val for ign in ignore_mst):
                                            data["Mã số thuế"] = mst_val
                                            print(f"  [HD-OCR] Found MST: {mst_val}")
                                            break
                        finally:
                            if not isinstance(pdf_source, str):
                                os.unlink(hd_path)
                    except Exception as e:
                        print(f"  [HD-OCR] Error: {e}")
                
                # Auto-classify based on content
                ocr_lower = ocr_text.lower()
                print(f"  OCR text contains 'petrolimex': {'petrolimex' in ocr_lower}")
                print(f"  OCR text contains 'xăng': {'xăng' in ocr_lower}")
                if any(x in ocr_lower for x in ['petrolimex', 'xăng', 'ron 95', 'ron95', 'diesel', 'dầu diesel']):
                    data["Phân loại"] = "Xăng xe"
                    print(f"  Classification set to: Xăng xe")
                elif any(x in ocr_lower for x in ['khách sạn', 'hotel', 'phòng nghỉ']):
                    data["Phân loại"] = "Dịch vụ phòng nghỉ"
                elif any(x in ocr_lower for x in ['nhà hàng', 'quán ăn', 'món ăn']):
                    data["Phân loại"] = "Dịch vụ ăn uống"
                else:
                    data["Phân loại"] = "Khác"
                print(f"  Final data: {data}")
                return data, []  # Return early for OCR path
            else:
                print(f"  OCR also failed for: {filename}")
                return data, []
        
        # Normalize newlines
        full_text = full_text.replace('\r\n', '\n').replace('\r', '\n')
        
        # CLEANUP: Remove garbage lines (e.g. debug JSON pointers like {'name': ...}) 
        clean_lines = []
        for line in full_text.split('\n'):
            line_strip = line.strip()
            
            # AGGRESSIVE CLEANUP for PSD.pdf garbage: 0'}2,950,000'}'}
            # FIRST: Check if this garbage contains a hidden number (Total)
            if "2,950,000" in line:
                 print(f"DEBUG_PSD_LINE: {repr(line)}")
            
            # Pattern: 0'}2,950,000'}'}
            # Try looser pattern: 0'}...digits...'}'}
            garbage_match = re.search(r"0'}([\d\.,]+)'\}'\}", line)
            if garbage_match:
                 val = garbage_match.group(1)
                 print(f"DEBUG: Found hidden total in garbage: {val}")
                 # Store it in data immediately
                 if not data["Số tiền sau"]:
                      data["Số tiền sau"] = val
                 # If we found it, we can strip the garbage wrapper but keep the number?
                 # Or just strip it all if we saved it?
                 # Let's keep the number in text just in case regexes need it
                 line = line.replace(garbage_match.group(0), " " + val + " ")

            # Repeatedly remove the garbage tokens until gone
            for _ in range(3):
                line = line.replace("0'}", "").replace("'}'}", "").replace("'}", "").replace("{'", "")
            
            # Filter lines that appear to be purely programming code/garbage
            # (Only filter if we failed to parse it as valid item above)
            if line_strip.startswith('{') or (line_strip.startswith("'") and line_strip.endswith("'")):
                 continue
            
            # Sanitization: Remove soft hyphens and null bytes
            line = line.replace('\xad', '').replace('\x00', '')
            clean_lines.append(line)
        full_text = "\n".join(clean_lines)

        
        # Fallback only works if we have a local file path
        if not full_text and isinstance(pdf_source, str):
            print(f"  Empty PDF text, checking for fallback text file...")
            base_name = os.path.splitext(os.path.basename(pdf_source))[0]
            folder = os.path.dirname(pdf_source)
            # Find closest matching text file (e.g. filename_00001.txt)
            for f in os.listdir(folder):
                # Check for files starting with the base name (ignoring the (1) vs (1)_0001 differences sometimes)
                # Simple check: startswith base_name and ends with .txt
                if f.startswith(base_name) and f.lower().endswith('.txt') and not f.startswith('debug_'):
                    txt_path = os.path.join(folder, f)
                    print(f"  -> Found fallback text file: {f}")
                    try:
                        # Try UTF-8 first
                        with open(txt_path, 'r', encoding='utf-8') as tf:
                            full_text = tf.read()
                        break
                    except UnicodeDecodeError:
                        try:
                             # Try CP1252 / ANSI
                             with open(txt_path, 'r', encoding='cp1252') as tf:
                                full_text = tf.read()
                             break
                        except Exception as e:
                             print(f"  -> Error reading fallback file (encoding): {e}")
                    except Exception as e:
                        print(f"  -> Error reading fallback file: {e}")

        if not full_text.strip():
            print(f"  Could not extract text (scanned PDF?): {filename}")
            # Set all fields to "không nhận diện được"
            for key in data:
                if key != "Tên file":
                    data[key] = "không nhận diện được"
            return data, []  # Return empty line_items

        
        # Extract services from text
        services = extract_services_from_text(full_text)
        
        # ============ EXTRACT FIELDS WITH MULTIPLE PATTERNS ============
        
        # Date extraction - try multiple patterns
        date_patterns = [
            r'Ngày\s*(\d{1,2})\s*tháng\s*(\d{1,2})\s*năm\s*(\d{4})',
            r'Ngày\s*(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{4})',
            r'(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{4})',
            # Multiline date matching with flexible noise skipping
            # Allows up to 100 chars of any text (including newlines) between parts
            r'Ngày[:\s]*(\d{1,2})[\s\S]{0,100}tháng[:\s]*(\d{1,2})[\s\S]{0,100}năm[:\s]*(\d{4})',
            # Bilingual Date: Ngày (day) 19 tháng (month) 12 năm (year) 2025
            r'Ngày(?:[^0-9]{0,35})?(\d{1,2})[\s\S]{0,35}tháng(?:[^0-9]{0,35})?(\d{1,2})[\s\S]{0,35}năm(?:[^0-9]{0,35})?(\d{4})'
        ]
        
        # Pre-process text for multiline date matching: remove newlines around Date keywords
        # This helps with: "Ngày 07 tháng 01\n năm 2026" -> "Ngày 07 tháng 01 năm 2026"
        minified_text = re.sub(r'(Ngày|tháng|năm)\s*\n\s*', r'\1 ', full_text, flags=re.IGNORECASE)
        
        for pattern in date_patterns:
            # Use DOTALL for multiline matching where needed, or standard for single line
            match = re.search(pattern, full_text, re.IGNORECASE | re.DOTALL)
            if match:
                day, month, year = match.groups()
                data["Ngày hóa đơn"] = f"{int(day):02d}/{int(month):02d}/{year}"
                break
        
        # SELLER TAX CODE (MST)
        # Strategy: 
        # 1. Look for MST explicitly associated with "Seller" or "Don vi ban"
        # 2. Look for MST generally but skip known "Provider" MSTs
        # 3. Handle spaces in MST (0 3 0 ...)
        
        # Known Provider MSTs to ignore (VNPT, Viettel, BKAV, etc often appear in footer)
        # 0106869738: VNPT, 0101360697: BKAV
        ignore_mst = ['0106869738', '0100684378', '0101245171', '0305482862', '0103243195', '0101360697']
        
        # Priority 0: Spaced MST Pattern (e.g. "0 3 0 1 4 3 3 9 8 4")
        # This is almost always the distinct Main Company MST at the header.
        # Must match sequence of digits separated by single spaces, length >= 10 digits
        spaced_mst_match = re.search(r'Mã số thuế[:\s]*((?:\d\s+){9,}[\d\s-]*\d)', full_text, re.IGNORECASE)
        if spaced_mst_match:
             potential_mst = spaced_mst_match.group(1).replace(' ', '').strip()
             if not any(x in potential_mst for x in ignore_mst):
                 data["Mã số thuế"] = potential_mst
                 print(f"  [MST] Found via Spaced pattern: {potential_mst}")
             else:
                 print(f"  [MST] Spaced match ignored: {potential_mst}")
        
        # Priority 0.5: VAT Code pattern (hotel invoices at footer)
        # Pattern: "VAT Code: 0300659964" or "VATCode: ..."
        if not data["Mã số thuế"]:
            vat_code_match = re.search(r'VAT\s*Code[:\s]*(\d{10,14})', full_text, re.IGNORECASE)
            if vat_code_match:
                potential_mst = vat_code_match.group(1).strip()
                if not any(x in potential_mst for x in ignore_mst):
                    data["Mã số thuế"] = potential_mst
        
        # Priority 1: Contextual match near "Đơn vị bán" or "Seller"
        # Only run if Priority 0 didn't find anything
        if not data["Mã số thuế"]:
             # Search in a window of text
             seller_block_match = re.search(r'(?:Đơn vị bán|Người bán|Seller)[^:]*[:\s]+(.*?)(?:Mã số thuế|MST|Tax code)[^:]*[:\s]*([0-9\s-]+)', full_text, re.IGNORECASE | re.DOTALL)
             if seller_block_match:
                  potential_mst = seller_block_match.group(2).replace(' ', '').strip()
                  # Check if it's a valid length MST
                  if len(potential_mst) >= 10 and not any(x in potential_mst for x in ignore_mst):
                      data["Mã số thuế"] = potential_mst
             else:
                  pass

        # Priority 2: Standard MST search if Priority 1 failed found nothing or ignored
        if not data["Mã số thuế"]:
            # Find ALL MSTs, then filter
            # Matches: "Mã số thuế: 030...", "MST: 030...", "Tax code: 030..."
            # Also handles spaced MST: "0 3 0 ..."
            all_mst_matches = re.finditer(r'(?:Mã số thuế|MST|Tax code)[^:]*[:\s]*([0-9\s-]+)', full_text, re.IGNORECASE)
            
            candidates = []
            for m in all_mst_matches:
                raw_mst = m.group(1).replace(' ', '').strip()
                # Clean trailing chars usually adhering to MST like -001 or just junk
                # Valid MST is usually 10-14 digits/chars
                clean_mst = re.match(r'[\d-]+', raw_mst)
                if clean_mst:
                    val = clean_mst.group(0)
                    # Check against ignore list: if ANY ignore_mst is a substring of val, OR val is substring of ignore_mst
                    is_ignored = any(ign in val for ign in ignore_mst) or any(val in ign for ign in ignore_mst)
                    
                    if 9 <= len(val) <= 14 and not is_ignored:
                         candidates.append(val)
            
            if candidates:
                # If multiple candidates, usually the FIRST one is the seller (top of page), 
                # unless the provider stamp is at the very top. 
                # But typically Seller info is top-left or top-center.
                
                # CONTEXT CHECK: If the line containing MST has "Gi?i ph?p", "Ph?m m?m", "Provider", ignore it
                final_candidates = []
                for c in candidates:
                     is_bad_context = False
                     # Find original match line to check context
                     for line in full_text.split('\n'):
                         if c in line.replace(' ', ''): # Approximation
                             if any(kw in line.lower() for kw in ['giải pháp', 'phần mềm', 'cung cấp bởi', 'phát hành bởi', 'created by', 'signature', 'ký bởi', 'bkav', 'ehoadon']):
                                 is_bad_context = True
                                 print(f"  [MST] Ignored candidate {c} due to bad context line: {line.strip()}")
                                 break
                     if not is_bad_context:
                         final_candidates.append(c)
                
                if final_candidates:
                    data["Mã số thuế"] = final_candidates[0]
                    print(f"  [MST] Found via Priority 2 (Standard): {data['Mã số thuế']}")
        
        # INVOICE NUMBER - Multiple patterns (order matters - more specific first)
        inv_patterns = [
            (r'(\d{8})\nSố HĐ\s*/\s*Invoice No\.', 0),  # C26MAP reverse: 00001348\nSố HĐ / Invoice No.:
            (r'(\d{4,8})\n\s*Số\s*\(?No\.?\)?[:\s]*', 0),  # 1C26MTA reverse: 4699\nSố (No.):
            (r'Số HĐ\s*/\s*Invoice No\.?[:\s]*[\n\s]*(\d{5,})', re.DOTALL),  # C26MAP: Số HĐ / Invoice No.:\n00001348
            (r'\(\s*VAT\s*INVOICE\s*\)[:\s]*(\d+)', 0),  # Special case: (VAT INVOICE) 00000043
            (r'Invoice No\.?[:\s]*[\n\s]*(\d{5,})', re.DOTALL),  # Generic Invoice No: 00001348
            (r'[Ss][ốo]\s*[(/]?\s*No\.?\s*[)/]?[:\s]*(\d{5,})', 0),  # M-INVOICE: Số(No.): 00007155
            r'[Ss][ốo][/\s]*[(]?\s*Invoice No\.?\s*[)]?[:\s]*(\d+)',
            r'\(RESTAURANT BILL\)\s*(\d+)',  # VNPT Restaurant: (RESTAURANT BILL) 00004501
            r'Số:\s*(\d+)',  # Explicit colon: Số: 00007155
            r'Số hóa đơn[:\s]*(\d+)',
            r'[Ss][ốo]\s*[(/]?\s*No\.?\s*[)/]?[:\s]*(\d+)',  # Catch-all: Số (No.)
            r's[éèẹẽe][: ]+\s*(\d+)',  # OCR typo: sé (Petrolimex)
            r'S[óố][: ]+\s*(\d+)',  # OCR typo: Só/Số
            # NOTE: Removed generic 'Số[:\s]+(\d+)' - too broad, matches addresses
        ]
        
        for pattern_item in inv_patterns:
            if isinstance(pattern_item, tuple):
                pattern, flags = pattern_item
                matches = list(re.finditer(pattern, full_text, re.IGNORECASE | flags))
            else:
                matches = list(re.finditer(pattern_item, full_text, re.IGNORECASE))
            
            for match in matches:
                num = match.group(1)
                
                # Context Check: Ensure it's not part of an address
                # Find the line containing this match
                start_pos, end_pos = match.span()
                line_start = full_text.rfind('\n', 0, start_pos) + 1
                line_end = full_text.find('\n', end_pos)
                if line_end == -1: line_end = len(full_text)
                line_content = full_text[line_start:line_end].lower()
                
                # Address keywords to avoid
                if any(kw in line_content for kw in ['đường', 'phố', 'phường', 'quận', 'thành phố', 'district', 'ward', 'street', 'thửa đất', 'tờ bản đồ', 'ấp ', 'xã ', 'tỉnh']):
                    # Only skip if it matches the generic "Số:" patterns (which are short/risky)
                    # If it explicitly says "Số HĐ" or "Invoice No", we might trust it more, 
                    # but even then "Số HĐ ... Đường..." is unlikely.
                    # Safe bet: If line looks like address, skip.
                    print(f"  [Invoice No] Skipped address-like match: '{match.group(0)}' in line '{line_content.strip()}'")
                    continue
                
                # Check for other junk matches
                if len(num) < 3: # Too short to be invoice number usually
                     continue
                     
                data["Số hóa đơn"] = num
                break
            
            if data["Số hóa đơn"]:
                break
        
        # Fallback: Extract from filename if missing
        if not data["Số hóa đơn"]:
            # Try to find a long number in filename?
            fname_for_num = filename if filename else (os.path.basename(pdf_source) if isinstance(pdf_source, str) else "")
            fname = os.path.splitext(fname_for_num)[0]
            # Split by underscores or hyphens
            parts = re.split(r'[_\-\s]', fname)
            # Filter for pure digit sequences, reasonable length (e.g. >3)
            # Avoid parts that look like dates if possible, but simplest is last long number
            nums = [p for p in parts if p.isdigit() and len(p) > 2]
            if nums:
                 data["Số hóa đơn"] = nums[-1] # Take the last number found pattern often has invoice num at end

        seller_patterns = [
            r'Đơn vị bán hàng\s*\([Ss]eller\)[:\s]*(.+)',  # M-INVOICE format
            r'Đơn vị bán\s*\([Ss]eller\)[:\s]*(.+)', # Standard
            r'Đơn vị bán\s*\(Seller\)[:\s]*(.+)',  # Specific exact match
            r'Tên người bán\s*\([Ss]eller\)[:\s]*(.+)',  # VNPT format
            r'Đơn vị bán hàng\s*\([Cc]ompany\)[:\s]*(.+)',  # MISA variation
            r'Đơn vị bán hàng[:\s]*(.+)',  # Simple format (Petrolimex)
            r'Tên đơn vị bán hàng[:\s]*(.+)',
            r'HỘ KINH DOANH[:\s]*(.+)',
            r'QUÁN[:\s]*(.+)',
            # NOTE: Removed 'Người bán' pattern - it captures 'Người bán hàng(Seller)' incorrectly
        ]
        for pattern in seller_patterns:
            match = re.search(pattern, full_text)
            if match:
                seller = match.group(1).strip()
                
                # Check for multi-line split (common in VNPT)
                # e.g. "CHI NHÁNH... (LOẠI HÌNH DOANH NGHIỆP:\nCÔNG TY TNHH)..."
                # Find start and end index of this match in full_text
                start_idx = match.end(1)
                # Look ahead for next line
                rest_of_text = full_text[match.end():]
                next_line_match = re.match(r'\n([^\n]+)', rest_of_text)
                if next_line_match:
                    next_line = next_line_match.group(1).strip()
                    # Heuristic to merge:
                    # 1. Seller line ends with ':', '(', or "DOANH NGHIỆP"
                    # 2. Next line starts with "CÔNG TY", "TẬP ĐOÀN", ")"
                    if (seller.endswith(':') or seller.endswith('(') or 'DOANH NGHIỆP' in seller[-15:]):
                         seller = seller + " " + next_line
                    
                # Clean up - remove (Seller): prefix and other text
                # Normalize newlines to spaces just in case
                seller = seller.replace('\n', ' ')
                
                # Robust cleanup of "Seller" / "Company" prefixes
                # Removes: "(Seller):", "Seller :", "(Company):", "Doanh nghiệp:", etc.
                seller = re.sub(r'^\s*[\(\[]?\s*(?:Seller|Company|Người bán|Doanh nghiệp|Tên đơn vị|Đơn vị bán)\s*[\)\]]?\s*[:\.\-]?\s*', '', seller, flags=re.IGNORECASE)
                seller = re.sub(r'^\s*\(?Issued\)?\s*[:\.\-]\s*', '', seller, flags=re.IGNORECASE) # Fix for (Issued) :
                seller = re.sub(r'^\s*[:\.\-]+\s*', '', seller) # Clean remaining colons/dashes
                seller = re.sub(r'\s*Mã số thuế.*$', '', seller, flags=re.IGNORECASE)
                seller = re.sub(r'\s*MST.*$', '', seller, flags=re.IGNORECASE)
                seller = re.sub(r'\s*Địa chỉ.*$', '', seller, flags=re.IGNORECASE)
                
                # Check for invalid seller content (captured footer text/codes)
                # Added 'địa chỉ', 'address' to prevent grabbing Address line
                if any(x in seller.lower() for x in ['mã nhận hóa đơn', 'code for checking', 'tra cứu tại', 'địa chỉ', 'address']):
                    continue
                
                # Check for placeholder capture
                if seller.lower().replace(':', '').strip() in ['(seller)', 'seller', 'người bán', 'tên đơn vị', '(buyer)', 'buyer', 'người mua']:
                    continue
                
                # 612167 fix: Reject page labels like "Trang 1/1", "Page 1"
                if re.match(r'^(Trang|Page)\s+\d', seller.strip(), re.IGNORECASE):
                    continue
                
                # Tiepkhach fix: Reject if seller is just a parenthetical label like "(Buyer)", "(Seller)"
                if re.match(r'^\s*\([A-Za-z\s]+\)\s*$', seller.strip()):
                    continue
                    
                if len(seller) > 5 or (len(seller) > 3 and 'QUÁN' in seller.upper()):
                    data["Đơn vị bán"] = seller
                    break
        
        # PRIORITY FALLBACK 1: First line(s) before first "Mã số thuế" - this is most reliable for MISA invoices
        # where seller company name is at the very top of the document
        if not data["Đơn vị bán"]:
            # Find the position of first "Mã số thuế" OR "MST"
            mst_pos = full_text.find("Mã số thuế")
            if mst_pos == -1:
                 mst_pos = full_text.find("MST")
                 
            if mst_pos > 0:
                # Get text before first MST
                text_before_mst = full_text[:mst_pos].strip()
                lines_before_mst = [l.strip() for l in text_before_mst.split('\n') if l.strip()]
                
                # First non-empty line that looks like a company name
                for idx, line in enumerate(lines_before_mst[:6]):  # Check first 6 lines to handle headers
                    # Must contain company keywords AND be reasonably long
                    if len(line) > 10 and any(kw in line.upper() for kw in ['CÔNG TY', 'TẬP ĐOÀN', 'CHI NHÁNH', 'NHÀ HÀNG', 'DNTN', 'HỘ KINH DOANH', 'QUÁN']):
                        # HDDT fix: If line has company keyword AND "HÓA ĐƠN" fused together,
                        # strip "HÓA ĐƠN" and everything after it instead of skipping
                        line_upper = line.upper()
                        has_hoadon = 'HÓA ĐƠN' in line_upper
                        # Exclude headers and BUYER info (but handle HÓA ĐƠN specially)
                        bad_keywords_no_hoadon = ['CỘNG HÒA', 'ĐỘC LẬP', 'TÊN NGƯỜI MUA', 'TÊN ĐƠN VỊ:', 'PHÂN PHỐI TỔNG HỢP DẦU KHÍ', 'ĐÃ ĐƯỢC KÝ ĐIỆN TỬ']
                        if any(bad in line_upper for bad in bad_keywords_no_hoadon):
                            continue
                        
                        # If line has "HÓA ĐƠN" fused with company name, strip it
                        if has_hoadon:
                            # Strip "HÓA ĐƠN" and everything after from the line
                            hoadon_pos = line_upper.find('HÓA ĐƠN')
                            line = line[:hoadon_pos].strip()
                            if len(line) <= 5:  # After stripping, too short = skip
                                continue
                        
                        if True:  # Replaces old "if not any(bad...)" block - keep indentation
                            # Multi-line company name: if next line is also uppercase text, merge
                            if idx + 1 < len(lines_before_mst):
                                next_line = lines_before_mst[idx + 1]
                                # Merge if next line doesn't contain MST markers and is short uppercase
                                if next_line and 'Mã số' not in next_line and 'Địa chỉ' not in next_line:
                                    if (next_line.isupper() or (len(next_line) < 40 and ':' not in next_line)) and 'PHÂN PHỐI' not in next_line.upper():
                                        line = line + " " + next_line
                            data["Đơn vị bán"] = line
                            break

        # FALLBACK 2: Ký bởi (Signed by) - common in footer, company name may span multiple lines
        if not data["Đơn vị bán"]:
            # Try multi-line pattern first: "Ký bởi:CÔNG TY...\nTHẾ THÊM"
            sign_match = re.search(r'(?:Ký bởi|Được ký bởi)[:\s]*([A-ZĐ][A-ZĐÀÁẢÃẠ\s]+(?:\n[A-ZĐÀÁẢÃẠ\s]+)?)', full_text)
            if sign_match:
                signer = sign_match.group(1).replace('\n', ' ').strip()
                # Only accept if it looks like a company name
                if len(signer) > 5 and any(x in signer.upper() for x in ['CÔNG TY', 'TẬP ĐOÀN', 'CHI NHÁNH', 'NHÀ HÀNG', 'DNTN']):
                    if not any(x in signer.lower() for x in ['địa chỉ', 'address', 'mã số', 'đã được ký']):
                        data["Đơn vị bán"] = signer
        
        # FALLBACK 3: Bottom Scan (Last 20 lines) - for Park Hyatt / Hotels
        if not data["Đơn vị bán"]:
            lines = [l.strip() for l in full_text.split('\n') if l.strip()]
            # Check last 20 lines
            for line in lines[-20:]:
                if len(line) > 5 and any(kw in line.upper() for kw in ['CÔNG TY', 'TẬP ĐOÀN', 'CHI NHÁNH', 'DNTN', 'HỘ KINH DOANH', 'HOTEL', 'KHÁCH SẠN', 'QUÁN']):
                    # Must be uppercase or mostly uppercase for Company Name
                    if line.isupper() or 'CÔNG TY' in line.upper() or 'QUÁN' in line.upper():
                         # Exclude headers/footer noise
                         if not any(bad in line.upper() for bad in ['HÓA ĐƠN', 'TRANG', 'PAGE', 'KÝ BỞI', 'GIẢI PHÁP', 'CUNG CẤP', 'ĐỊA CHỈ', 'MST:', 'VAT CODE']):
                              data["Đơn vị bán"] = line
                              break

        # Mã CQT (Standard PDF)
        # Matches: "Mã của cơ quan thuế: ...", "Mã CQT: ..."
        cqt_match = re.search(r'(?:Mã|Ma)\s*(?:của)?\s*(?:CQ|cơ\s*quan)\s*thuế[:\s]*([A-Z0-9\-]+)', full_text, re.IGNORECASE)
        if cqt_match:
            data["Mã CQT"] = cqt_match.group(1)
            
        # SERIAL NUMBER (Ký hiệu) - Multiple patterns INCLUDING "Series"
        serial_patterns = [
            r'Ký hiệu\s*/\s*Serial[:\s]*([A-Z0-9]+)',  # C26MAP: Ký hiệu / Serial: 1C26MAP
            r'[KK]ý hiệu\s*/\s*\([Ss]erial(?:\s*No\.?)?\)[:\s]*([A-Z0-9]+)',  # Format with slash: Ký hiệu/ (Serial No)
            r'[KK]ý hiệu\s*\([Ss]erial\)[:\s]*([A-Z0-9]+)',  # VNPT: Ký hiệu(Serial): 1K25THA
            r'[KK]ý hiệu\s*\([Ss]erial(?:\s*No\.?)?\)[:\s]*([A-Z0-9]+)',  # M-INVOICE
            r'[KK]ý hiệu\s*\([Ss]eries\)[:\s]*([A-Z0-9]+)',  # VNPT uses "Series"
            r'[KK]ý hiệu[:\s]*([A-Z0-9]+)',
            r'Mẫu số\s*-\s*[KK]ý hiệu[^:]*[:\s]*([A-Z0-9]+)',
        ]
        for pattern in serial_patterns:
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                data["Ký hiệu"] = match.group(1)
                break
        
        # SECURITY CODE (Mã tra cứu) - Multiple patterns
        security_patterns = [
            r'Mã tra cứu hoá đơn[:\s]*([A-Za-z0-9]+)',  # C26MAP: Mã tra cứu hoá đơn: 9751Opera19012026
            r'Mã nhận hóa đơn\s*\([Cc]ode for checking\)[:\s]*([A-Z0-9]+)',  # Special case
            r'Mã nhận hóa đơn[:\s]*([A-Za-z0-9]+)',  # Simple "Mã nhận hóa đơn: 5c57d33"
            r'Mã tra cứu\s*\([Ll]ookup\s*code\)[:\s]*([A-Za-z0-9_]+)',  # VNPT: Mã tra cứu(Lookup code):HCM...
            r'Mã tra cứu hóa đơn\s*\([Ii]nvoice code\)[:\s]*([A-Za-z0-9_]+)',  # MISA variation
            r'Mã tra cứu(?:\s*HĐĐT)?(?:\s*này)?[:\s]*([A-Za-z0-9_]+)',
            r'Mã tra cứu\(Invoice code\)[:\s]*([A-Za-z0-9_]+)',  # MISA no-space
            r'Mã số bí mật[:\s]*([A-Za-z0-9_]+)',
            r'Security Code\)[:\s]*([A-Z0-9]+)',
            r'Mã tra cứu[:\s]*([A-Za-z0-9]+)',
            r'[Ll]ookup\s*code[):\s]*([A-Za-z0-9]+)',
            r'Ma tra cuu[:\s]*([A-Za-z0-9]+)', # Non-accented
            r'Mã tra cứu\s*\([Cc]ode\)[:\s]*([A-Za-z0-9]+)', # K26THT: Mã tra cứu (Code): ...
            r'với mã[:\s]*([A-Za-z0-9]+)', # C26MCX: lấy hóa đơn với mã: ...
            r'nhập mã\s+([A-Za-z0-9]+)', # NEW: "nhập mã [CODE] để lấy hóa đơn"
            r'provided code[^:]*[:\s]*([A-Za-z0-9]+)', # NEW: "provided code to get invoice: [CODE]"
        ]
        for pattern in security_patterns:
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                code = match.group(1)
                # Avoid capturing URL parts or headers as code
                if any(x in code.lower() for x in ["http", "tracuu", "website", "invoice", "check", ".com", ".vn", "please", "vui lòng", "quý khách", "access"]):
                    continue

                # Allow longer codes (VNPT uses 32 chars)
                if 5 <= len(code) <= 35:
                    data["Mã tra cứu"] = code
                    break
                elif len(code) > 35:
                    # Very long code might be CQT, store separately
                    if not data["Mã CQT"]:
                        data["Mã CQT"] = code
        
        # Fallback: If no Lookup Code found but CQT Code exists (often treated as the unique ID for VNPT/others)
        # Check this AFTER extracting Code to avoid overwriting invalid long codes
        
        # Fallback for PSD.pdf where "Mã tra cứu" is not clearly labeled but looks like a long code
        if not data["Mã tra cứu"]:
             # Look for long string of mixed Upper/Digits in footer area (last 200 chars)
             footer_text = full_text[-500:] 
             # Common format: no label, just the code
             potentials = re.findall(r'\b[A-F0-9]{8,}\b', footer_text)
             for p in potentials:
                 if len(p) >= 10 and not p.isdigit(): # Mix of chars, likely Lookup Code
                     if "0100" not in p and "030" not in p: # Avoid tax codes
                         # HDDT fix: Skip codes on lines with "Serial number" or "Ký điện tử" context
                         is_serial = False
                         for ft_line in footer_text.split('\n'):
                             if p in ft_line:
                                 if any(kw in ft_line.lower() for kw in ['serial number', 'serial no', 'ký điện tử', 'ký điện tư', 'chữ ký số']):
                                     is_serial = True
                                 break
                         if is_serial:
                             continue
                         data["Mã tra cứu"] = p
                         break
        
        # TAX CODE (MST đơn vị bán) - Look for seller's tax code (first one)
        tax_patterns = [
            r'Mã số thuế\s*\([Tt]ax\s*code\)[:\s]*([\d\-\u00AD\s]+)',  # Added \s for spaced numbers
            r'(?:MST|Mã số thuế)[/\s]*\([Tt]ax [Cc]ode\)[:\s]*([\d\-\u00AD\s]+)',
            r'MST/CCCD[^:]*[:\s]*([\d\-\u00AD\s]+)',
            r'(?:MST|Mã số thuế)[:\s]*([\d\-\u00AD\s]+)',
        ]
        tax_codes = []
        for pattern in tax_patterns:
            matches = re.findall(pattern, full_text, re.IGNORECASE)
            # Clean up matches - remove soft hyphens AND spaces
            cleaned_matches = []
            for m in matches:
                clean = m.replace('\u00AD', '').replace(' ', '').strip()
                # Verify it looks like a tax code (at least 10 chars, digits/hyphens)
                if len(clean) >= 10 and any(c.isdigit() for c in clean):
                    cleaned_matches.append(clean)
            tax_codes.extend(cleaned_matches)
        
        if len(tax_codes) >= 1 and not data["Mã số thuế"]:
             # Filter ignore list
             valid_mst = [t for t in tax_codes if not any(ign in t for ign in ignore_mst) and not any(t in ign for ign in ignore_mst)]
             if valid_mst:
                 data["Mã số thuế"] = valid_mst[0]  # Seller's tax code (first one)
        
        # PETROLIMEX SPECIFIC: OCR often messes up "Ma so thue" label or merges it.
        # Look for "Ma so thue: 0300555450" or similar in OCR text (normalized)
        if not data["Mã số thuế"]:
             # Normalize text: lower, remove accents
             norm_text = full_text.lower().replace('á', 'a').replace('à', 'a').replace('ã', 'a').replace('ạ', 'a').replace('ả', 'a') \
                                          .replace('é', 'e').replace('è', 'e').replace('ẽ', 'e').replace('ẹ', 'e').replace('ẻ', 'e') \
                                          .replace('ô', 'o').replace('ố', 'o').replace('ồ', 'o').replace('ỗ', 'o').replace('ộ', 'o').replace('ổ', 'o') \
                                          .replace('ê', 'e').replace('ế', 'e').replace('ề', 'e').replace('ễ', 'e').replace('ệ', 'e').replace('ể', 'e')
             # Pattern: "ma so thue" or "ma se thue" (OCR typo) followed by digits
             # Handle "Ma sé thué" -> "ma se thue"
             petro_mst = re.search(r'(?:ma\s+s[eoc]\s+thue|ma\s+so\s+thue|ma\s+s.\s+thue|mst|tax code)[^0-9]*([0-9]{10,14})', norm_text)
             if petro_mst:
                 mst_cand = petro_mst.group(1)
                 if not any(x in mst_cand for x in ignore_mst):
                      data["Mã số thuế"] = mst_cand
        
        # CQT CODE - Multiple patterns (include soft hyphen \u00AD used in some PDFs)
        cqt_patterns = [
            r'Mã\s*(?:của\s*)?[Cc]ơ quan thuế[:\s]*([A-Za-z0-9\-\u00AD]+)',  # M-INVOICE
            r'Mã\s*(?:của\s*)?[Cc]ơ quan thuế\s*\([Tt]ax authority code\)[:\s]*([A-Za-z0-9\-\u00AD]+)',
            r'Mã\s*CQT\s*\([Cc]ode\)[:\s]*([A-Za-z0-9\-\u00AD]+)',
            r'Mã\s*CQT[:\s]*([A-Za-z0-9\-\u00AD]+)',
            r'MCQT\s*[:\s]+([A-Za-z0-9\-\u00AD]+)',  # Tiepkhach: "MCQT :M1-26-..."
            r'Tax authority code[:\s]*([A-Za-z0-9\-\u00AD]+)',
        ]
        for pattern in cqt_patterns:
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                # Replace soft hyphen with regular hyphen
                cqt_code = match.group(1).strip().replace('\u00AD', '-')
                data["Mã CQT"] = cqt_code
                break
        
        # FINAL FALLBACK: If Lookup Code is still empty, use CQT Code
        # REMOVED due to User Request: "không được lấy mã cơ quan thuế thay vào cho mã tra cứu"
        # if not data["Mã tra cứu"] and data["Mã CQT"]:
        #    data["Mã tra cứu"] = data["Mã CQT"]
        
        # LOOKUP LINK - Multiple patterns
        link_patterns = [
            r'Tra cứu hóa đơn tại\s*\([^)]+\)[:\s]*(https?://[^\s]+)',  # VNPT: Tra cứu hóa đơn tại (Lookup the invoice at):https://...
            r'Tra cứu hóa đơn tại[:\s]*(https?://[^\s]+)',  # Simple format
            r'(?:Tra cứu[^:]*tại|Trang tra cứu|website)[:\s]*(https?://[^\s]+)',
            r'(https?://[^\s]*(?:tracuu|tra-cuu|invoice|vnpt-invoice|minvoice|hddt)[^\s]*)',  # Added hddt for Thế Giới Di Động
            # HDDT fix: Generic https:// URL near invoice keywords ("hóa đơn", "tải về")
            r'(?:tải|lấy|xem|download)\s+(?:về\s+)?hóa đơn[^\n]*(https?://[^\s]+)',  # "tải về hóa đơn ... https://..."
            r'(https?://[^\s]+)[^\n]*(?:hóa đơn|tải về|lấy hóa đơn)',  # URL before "hóa đơn" keyword
            # Pattern for links without http/https (e.g. hoadon.pvoil.vn, tracuu.wininvoice.vn)
            r'(?:Tra cứu[^:]*tại|Trang tra cứu|website)[:\s]*([a-zA-Z0-9.-]+\.[a-zA-Z]{2,}(?:/[^\s]*)?)',
        ]
        for pattern in link_patterns:
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                link = match.group(1).rstrip('.').rstrip(',')
                # If link doesn't start with http, prepend http://
                if not link.lower().startswith('http') and not link.lower().startswith('www'):
                     link = "http://" + link
                if "www" in link.lower() and not link.lower().startswith('http'):
                     link = "http://" + link
                
                # Filter out junk that might be matched as domain
                if '.' in link and len(link) > 5:
                     data["Link lấy hóa đơn"] = link
                     break
        
        # USER REQUEST: "mã tra cứu luôn hiển thị bên cạnh link tra cứu"
        # Search specifically for Code near Link (using generic link patterns if exact link mismatch)
        if not data["Mã tra cứu"]:
             # Pattern: Link followed by Code (within proximity)
             # Matches: http://... <space> CODE
             # We iterate to find the Best candidate
             prox_patterns = [
                 r'(?:https?://[^\s]+)[\s\n]+([A-Za-z0-9]{6,50})\b', # Link -> Code
                 r'\b([A-Za-z0-9]{6,50})[\s\n]+(?:https?://[^\s]+)', # Code -> Link
             ]
             for p in prox_patterns:
                 matches = re.finditer(p, full_text, re.IGNORECASE)
                 for m in matches:
                      cand = m.group(1)
                      # Filter junk
                      if cand.lower() in ["website", "http", "https", "link", "tại", "vnbox", "vnpt", "invoice"]:
                          continue
                      if "tracuu" in cand.lower():
                          continue
                      # Filter if it matches MST
                      if data["Mã số thuế"] and cand == data["Mã số thuế"]:
                          continue
                      # Filter if it looks like a pure date (dd/mm/yyyy no separators? rare) or phone number
                      
                      # Validation: Lookup Codes usually complicated. 
                      # If purely numeric, risky? No, some are numeric.
                      
                      if 6 <= len(cand) <= 50:
                           data["Mã tra cứu"] = cand
                           break
                 if data["Mã tra cứu"]:
                     break
        
        # AMOUNTS - Multiple patterns
        # Before tax
        before_tax_patterns = [
            r'Cộng tiền hàng\s*/\s*Total charges[:\s]*([\d\.,]+)',  # C26MAP: Cộng tiền hàng / Total charges: 6.615.000
            r'Cộng tiền hàng[^:]*[:\s]*([\d\.,]+)',
            r'Cộng ti[êề]n hàng[^:]*[:\s]*([\d\.,]+)', # OCR typo: tiên
            r'Tổng tiền chưa thuế[^:]*[:\s]*([\d\.,]+)',  # M-INVOICE
            r'Thành ti[êềẫ]n trước thuế[^:]*[:\s]*([\d\.,]+)', # OCR typo: tiễn
            r'Amount before VAT[^:]*[:\s]*([\d\.,]+)',
            r'Sub total[^:]*[:\s]*([\d\.,]+)',
        ]
        for pattern in before_tax_patterns:
            matches = re.findall(pattern, full_text, re.IGNORECASE)
            if matches:
                # Take the LAST match as it's likely the grand total on the last page
                data["Số tiền trước Thuế"] = matches[-1]
                break
        
        # HDDT fix: "Tiền hàng:" multi-column format (e.g. "Tiền hàng: 0 0 0 0 185.185 0 185.185")
        # Take the last non-zero value from the row
        if not data["Số tiền trước Thuế"]:
            tien_hang_match = re.search(r'Tiền hàng[:\s]+([\d\s.,]+)', full_text, re.IGNORECASE)
            if tien_hang_match:
                nums_str = tien_hang_match.group(1).strip()
                nums_list = re.findall(r'[\d]+(?:[.,][\d]+)*', nums_str)
                # Take last non-zero value
                for n in reversed(nums_list):
                    if parse_money(n) > 0:
                        data["Số tiền trước Thuế"] = n
                        break
        
        # VAT AMOUNT (Tiền thuế)
        vat_patterns = [
            r'Tiền thuế GTGT\s*/\s*VAT[:\s]*([\d\.,]+)',  # C26MAP: Tiền thuế GTGT / VAT: 529.200
            r'Tổng tiền thuế GTGT \d+%[:\s]*([\d\.,]+)', # Specific rate line
            r'\|?Tiền thu[êế] GTGT\s*\(\s*\d+\s*%\s*\)\s*([\d\.,]+)', # MOST SPECIFIC: |Tiền thuê GTGT ( 8% ) 59.265
            r'\|?Tiền thu[êế] GTGT[^:]*[:\s]+(\d[\d\.,]+)', # |Tiền thuê GTGT: 59.265
            r'Tiền thuế\s*\(VAT\s*Amount\)[^:]*[:\s]*([\d\.,]+)',
            r'Tổng tiền thuế[^:]*[:\s]*([\d\.,]+)',  # M-INVOICE
            r'Tiền thu[êế] GTGT[^:]*[:\s]+(\d[\d\.,]+)', # OCR typo: thuê (ensure starts with digit)
            r'VAT amount[^:]*[:\s]*([\d\.,]+)',
            r'Cộng tiền thuế GTGT[^:]*[:\s]*([\d\.,]+)',
        ]
        for pattern in vat_patterns:
            matches = re.findall(pattern, full_text, re.IGNORECASE)
            if matches:
                 # Take the LAST match
                data["Tiền thuế"] = matches[-1]
                break
        
        # VAT RATE BREAKDOWN (detect amounts by specific tax rates)
        # Format 1 - Sapo: "Thuế suất 8% : 995,000 79,600 1,074,600" where 2nd number is tax
        # Format 2 - M-Invoice: "Tổng tiền chịu thuế suất: 8% 655.000 52.400 707.400" where 2nd number after % is tax
        # Format 3 - Simple: "Tổng tiền thuế GTGT 8%: 17.592,59"
        
        # Multi-column patterns (Sapo, M-Invoice, MISA): X% before_tax tax_amount total
        # NOTE: Use [^:\n]* instead of [^:]* to prevent matching across newlines
        multi_col_patterns = [
            # Sapo/MISA format: "Thuế suất 8%(VAT rate 8%): before tax total" - allow text after %
            (r'Thuế suất\s*0\s*%[^:\n]*[:\s]+(\d[\d\.,]*)\s+(\d[\d\.,]*)\s+(\d[\d\.,]*)', "Thuế 0%", 2),
            (r'Thuế suất\s*5\s*%[^:\n]*[:\s]+(\d[\d\.,]*)\s+(\d[\d\.,]*)\s+(\d[\d\.,]*)', "Thuế 5%", 2),
            (r'Thuế suất\s*8\s*%[^:\n]*[:\s]+(\d[\d\.,]*)\s+(\d[\d\.,]*)\s+(\d[\d\.,]*)', "Thuế 8%", 2),
            (r'Thuế suất\s*10\s*%[^:\n]*[:\s]+(\d[\d\.,]*)\s+(\d[\d\.,]*)\s+(\d[\d\.,]*)', "Thuế 10%", 2),
            # Golden Gate 5-column format: "thuế suất khác... 8% before discount after_disc TAX total"
            (r'Thuế suất\s*khác[^0-9\n]*8\s*%\s+(\d[\d\.,]*)\s+(\d[\d\.,]*)\s+(\d[\d\.,]*)\s+(\d[\d\.,]*)\s+(\d[\d\.,]*)', "Thuế 8%", 4),
            (r'Thuế suất\s*khác[^0-9\n]*10\s*%\s+(\d[\d\.,]*)\s+(\d[\d\.,]*)\s+(\d[\d\.,]*)\s+(\d[\d\.,]*)\s+(\d[\d\.,]*)', "Thuế 10%", 4),
            (r'Thuế suất\s*khác[^0-9\n]*5\s*%\s+(\d[\d\.,]*)\s+(\d[\d\.,]*)\s+(\d[\d\.,]*)\s+(\d[\d\.,]*)\s+(\d[\d\.,]*)', "Thuế 5%", 4),
            # M-Invoice format: "Tổng tiền chịu thuế suất... 8% before tax total"
            (r'Tổng tiền chịu thuế suất[^:\n]*:\s*0\s*%\s+(\d[\d\.,]*)\s+(\d[\d\.,]*)\s+(\d[\d\.,]*)', "Thuế 0%", 2),
            (r'Tổng tiền chịu thuế suất[^:\n]*:\s*5\s*%\s+(\d[\d\.,]*)\s+(\d[\d\.,]*)\s+(\d[\d\.,]*)', "Thuế 5%", 2),
            (r'Tổng tiền chịu thuế suất[^:\n]*:\s*8\s*%\s+(\d[\d\.,]*)\s+(\d[\d\.,]*)\s+(\d[\d\.,]*)', "Thuế 8%", 2),
            (r'Tổng tiền chịu thuế suất[^:\n]*:\s*10\s*%\s+(\d[\d\.,]*)\s+(\d[\d\.,]*)\s+(\d[\d\.,]*)', "Thuế 10%", 2),
            # Format: "Thuế suất GTGT: 8%" (followed by tax amount or just standalone)
            (r'Thuế suất(?:\s*GTGT)?[:\s]*8\s*%\s*Tiền thuế GTGT[:\s]*(\d[\d\.,]*)', "Thuế 8%", 1),
            (r'Thuế suất(?:\s*GTGT)?[:\s]*10\s*%\s*Tiền thuế GTGT[:\s]*(\d[\d\.,]*)', "Thuế 10%", 1),
            (r'Thuế suất(?:\s*GTGT)?[:\s]*5\s*%\s*Tiền thuế GTGT[:\s]*(\d[\d\.,]*)', "Thuế 5%", 1),
            # Format: "Tiền thuế GTGT: ( 8% ) 37.037" (C26MCX)
            (r'Tiền thuế GTGT[:\s]*[\(\[]?\s*8\s*%\s*[\)\]]?\s*(\d[\d\.,]*)', "Thuế 8%", 1),
            (r'Tiền thuế GTGT[:\s]*[\(\[]?\s*10\s*%\s*[\)\]]?\s*(\d[\d\.,]*)', "Thuế 10%", 1),
            (r'Tiền thuế GTGT[:\s]*[\(\[]?\s*5\s*%\s*[\)\]]?\s*(\d[\d\.,]*)', "Thuế 5%", 1),
            # Format: "Tiền thuế ( 10% ): 154.545" (C26TKM) - no GTGT
            (r'Tiền thuế[:\s]*[\(\[]?\s*10\s*%\s*[\)\]]?[:\s]*(\d[\d\.,]*)', "Thuế 10%", 1),
            (r'Tiền thuế[:\s]*[\(\[]?\s*8\s*%\s*[\)\]]?[:\s]*(\d[\d\.,]*)', "Thuế 8%", 1),
            (r'Tiền thuế[:\s]*[\(\[]?\s*5\s*%\s*[\)\]]?[:\s]*(\d[\d\.,]*)', "Thuế 5%", 1),
            # Loose format: "Tiền thuế ... 10% ... amount"
            (r'Tiền thuế[^%\d]*10\s*%.*?(\d[\d\.,]*)', "Thuế 10%", 1),
            (r'Tiền thuế[^%\d]*8\s*%.*?(\d[\d\.,]*)', "Thuế 8%", 1),
            (r'Tiền thuế[^%\d]*5\s*%.*?(\d[\d\.,]*)', "Thuế 5%", 1),
        ]
        for pattern, column, group_idx in multi_col_patterns:
            matches = list(re.finditer(pattern, full_text, re.IGNORECASE))
            if matches:
                data[column] = matches[-1].group(group_idx)
        
        # Single-value patterns (try if multi-column didn't find anything)
        if not any(data[c] for c in ["Thuế 0%", "Thuế 5%", "Thuế 8%", "Thuế 10%", "Thuế khác"]):
            simple_patterns = [
                # "Tổng tiền thuế GTGT 8%: 17.592,59" format
                (r'Tổng tiền thuế GTGT\s*0\s*%\s*[:\s]*(\d[\d\.,]*)', "Thuế 0%"),
                (r'Tổng tiền thuế GTGT\s*5\s*%\s*[:\s]*(\d[\d\.,]*)', "Thuế 5%"),
                (r'Tổng tiền thuế GTGT\s*8\s*%\s*[:\s]*(\d[\d\.,]*)', "Thuế 8%"),
                (r'Tổng tiền thuế GTGT\s*10\s*%\s*[:\s]*(\d[\d\.,]*)', "Thuế 10%"),
                # "Thuế GTGT (8%): amount" format
                (r'(?:thuế gtgt|VAT)\s*[\(\[]?\s*0\s*%\s*[\)\]]?\s*[:\s]*(\d[\d\.,]*)', "Thuế 0%"),
                (r'(?:thuế gtgt|VAT)\s*[\(\[]?\s*5\s*%\s*[\)\]]?\s*[:\s]*(\d[\d\.,]*)', "Thuế 5%"),
                (r'(?:thuế gtgt|VAT)\s*[\(\[]?\s*8\s*%\s*[\)\]]?\s*[:\s]*(\d[\d\.,]*)', "Thuế 8%"),
                (r'(?:thuế gtgt|VAT)\s*[\(\[]?\s*10\s*%\s*[\)\]]?\s*[:\s]*(\d[\d\.,]*)', "Thuế 10%"),
                # "Tiền thuế GTGT ( 8% ) amount" format (OCR typo)
                (r'Tiền thu[êế] GTGT\s*\(\s*0\s*%\s*\)\s*(\d[\d\.,]*)', "Thuế 0%"),
                (r'Tiền thu[êế] GTGT\s*\(\s*5\s*%\s*\)\s*(\d[\d\.,]*)', "Thuế 5%"),
                (r'Tiền thu[êế] GTGT\s*\(\s*8\s*%\s*\)\s*(\d[\d\.,]*)', "Thuế 8%"),
                (r'Tiền thu[êế] GTGT\s*\(\s*10\s*%\s*\)\s*(\d[\d\.,]*)', "Thuế 10%"),
            ]
            for pattern, column in simple_patterns:
                matches = re.findall(pattern, full_text, re.IGNORECASE)
                if matches:
                    data[column] = matches[-1]
                matches = re.findall(pattern, full_text, re.IGNORECASE)
                if matches:
                    data[column] = matches[-1]
                    
        # Extra Fallback: "Tiền thuế" with simple label (often found in Footer)
        if not data["Tiền thuế"]:
            # Try finding just loose "Tiền thuế ...."
            simple_tax = re.search(r'(?:Tiền thuế|Thuế GTGT|VAT)\s*[\(\d%]*\)?[:\s]*([0-9]+[.,][0-9]+)', full_text, re.IGNORECASE)
            if simple_tax:
                 data["Tiền thuế"] = simple_tax.group(1)
            
            # If still not found, try finding line with "10%" or "8%" and taking the number at the end
            if not data["Tiền thuế"]:
                 rate_lines = re.findall(r'(?:10%|8%)\s+([0-9][\d\.,]+)', full_text)
                 if rate_lines:
                     # Usually the last number on a "10%" line is the tax amount or total
                     # This is risky but better than nothing for 0318...pdf
                     pass

        # SERVICE CHARGE (Phí PV)
        # Pattern: Phí PV(Sevice change): 400.507
        pv_match = re.search(r'Phí\s*PV[^:]*[:\s]*([\d\.,]+)', full_text, re.IGNORECASE)
        if pv_match:
            data["Phí PV"] = pv_match.group(1)



        # If we found total tax but no breakdown, calculate rate from amounts
        if data["Tiền thuế"] and not any(data[c] for c in ["Thuế 0%", "Thuế 5%", "Thuế 8%", "Thuế 10%"]):
            total_tax = parse_money(data["Tiền thuế"])
            before_tax = parse_money(data["Số tiền trước Thuế"])
            if total_tax and before_tax and before_tax > 0:
                rate = round(total_tax / before_tax * 100)
                if rate in [0, 5, 8, 10]:
                    key = f"Thuế {rate}%"
                    # Calculate implicitly to ensure correct rounding if we are filling it
                    # But here we just move Total Tax to the bucket.
                    # However, if we ever needed to recalculate Pre-Tax, we need consistent logic.
                    data[key] = data["Tiền thuế"]
                else:
                    data["Thuế khác"] = data["Tiền thuế"]
        
        # Priority 3.5: Handle "Hóa đơn bán hàng" (Sales Invoice - direct sale, often no dedicated TAX line)
        # Identify by Title or "Total amount" pattern from log: "a, dịch vụ(Total amount): 5.400.000"
        is_sales_invoice = "HÓA ĐƠN BÁN HÀNG" in full_text.upper() or "(SALES INVOICE)" in full_text.upper()
        
        if is_sales_invoice:
             # Try to find total amount if missing
             if not data["Số tiền sau"]:
                 # Pattern from log: "a, dịch vụ(Total amount): 5.400.000"
                 # And generic "Total amount: ..."
                 sales_total_match = re.search(r'(?:Total amount|dịch vụ\s*\(Total amount\))[:\s]*([\d\.,]+)', full_text, re.IGNORECASE)
                 if sales_total_match:
                     data["Số tiền sau"] = sales_total_match.group(1).strip()
            
             # For Sales Invoice, if 'Tax' is missing, usually Header Amount = Total Amount
             if data["Số tiền sau"] and not data["Số tiền trước Thuế"]:
                 data["Số tiền trước Thuế"] = data["Số tiền sau"]
                 # Tax is implicitly included or 0, but usually we just leave Tax empty or 0
        
        # Priority 4: Auto-classify "Dịch vụ du lịch"
        # Check Seller Name for keywords
        seller_upper = data.get("Đơn vị bán", "").upper()
        full_text_upper = full_text.upper()
        
        if "DU LỊCH" in seller_upper or "TRAVEL" in seller_upper or "DỊCH VỤ DU LỊCH" in full_text_upper:
            data["Phân loại"] = "Dịch vụ du lịch"
        
        # Refine Seller Name for this specific invoice if it was cut off
        # Text: "HỘ KINH DOANH DỊCH VỤ DU LỊCH NHÂN LỢI PHÁT"
        if "NHÂN LỢI PHÁT" in full_text_upper and not data["Đơn vị bán"]:
             seller_match = re.search(r'HỘ KINH DOANH DỊCH VỤ DU LỊCH [^\n]+', full_text, re.IGNORECASE)
             if seller_match:
                 data["Đơn vị bán"] = seller_match.group(0).strip()
        
        # After tax (total payment)
        after_tax_patterns = [
            r'Tổng cộng\s*/\s*Total Amount[:\s]*([\d\.,]+)',  # C26MAP: Tổng cộng / Total Amount: 7.144.200
            # Golden Gate 5-column FIRST (most specific): 5 numbers separated by spaces
            r'Tổng cộng tiền thanh toán\s*\(Total amount\)\s*([\d\.,\s]+)', # Golden Gate: 5 numbers, take the last one
            # 4-column patterns (discount, before_tax, tax, total) - MUST be before 3-column
            r'Tổng cộng\s*\(Total amount\)\s*[:]\s*[\d\.,]+\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)',  # 1C26MDA: 0 1.224.000 97.920 1.321.920
            r'Tổng tiền chịu thuế suất\s*\(Total amount\)\s*[:]\s*\d+%\s+[\d\.,]+\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)',  # 1C26MDA tax rate line
            # 3-column patterns
            r'Tổng cộng\s*\(Total amount\)\s*[:]\s*([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)', # File 2025_812...
            r'Tổng\s*cộng\s*\([Tt]otal\)?[:\s]*([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)', # MISA: Tổng cộng(Total): 375.000 30.000 405.000
            r'Tổngcộng[:\s]*([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)', # No-space: Tổngcộng: 2.816.100 256.158 3.072.258
            r'Tổng cộng\s*[:]\s*([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)', # SAPO/EasyInvoice: Tổng cộng: [Before] [VAT] [Total]
            r'Tổng tiền chịu thuế suất.*[:\s]*[\d\.,]*%\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)', # M-INVOICE table summary
            r'[Tt]ổng\s*tiền\s*thanh\s*toán\s*\([^)]+\)[:\s]*([\d\.,]+)',  # MISA: Tổng tiền thanh toán (Total amount): 1.800.000
            r'[IT].{1,3}ng\s*số\s*ti[êề]n\s*thanh\s*toán[:\s]*([\d\.,]+)', # OCR typo: Iông, tiên (Petrolimex)
            r'Cộng tiền hàng hóa, dịch vụ[:\s]*[\d\.,]+\s+[\d\.,]+\s+([\d\.,]+)', # File 1226-TK-200k (Total on next line)
            r'[Tt]ổng\s*cộng\s*tiền\s*thanh\s*toán[^:]*[:\s]*([\d\.,]+)',
            r'[Tt]otal\s*payment[^:]*[:\s]*([\d\.,]+)',
            r'TỔNG CỘNG TIỀN THANH TOÁN[^:]*[:\s]*([\d\.,]+)',
            r'Tổng cộng[:\s]+([\d\.,]+)\s+[\d\.,]+\s+([\d\.,]+)',  # Multi-page format
            r'thuế suất:\s*\d+%\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)',  # Tax rate line
        ]
        for pattern in after_tax_patterns:
            matches = list(re.finditer(pattern, full_text, re.IGNORECASE))
            if matches:
                match = matches[-1] # Take the LAST match
                
                # Handle multi-column format (before_tax, vat, after_tax)
                if match.lastindex and match.lastindex >= 3:
                    # Specific check for SAPO/EasyInvoice where Group 1=Before, Group 2=VAT, Group 3=Total
                    # Overwrite existing values as summary line is more reliable
                    data["Số tiền sau"] = match.group(match.lastindex)
                    data["Tiền thuế"] = match.group(2)
                    data["Số tiền trước Thuế"] = match.group(1)
                elif match.lastindex and match.lastindex >= 2:
                    data["Số tiền sau"] = match.group(match.lastindex)
                    # For format with 2 columns, be careful about overwriting
                    if not data["Số tiền trước Thuế"]:
                        data["Số tiền trước Thuế"] = match.group(1)
                else:
                    # Single group or Golden Gate complex case
                    val = match.group(1)
                    # Use helper split if it looks like multiple numbers (Golden Gate)
                    parts = val.strip().split()
                    if len(parts) >= 5 and all(c in '0123456789.,' for c in ''.join(parts)):
                         # Golden Gate 5-column: before discount after_disc TAX total
                         # 1.656.000 100.000 1.556.000 124.480 1.680.480
                         data["Số tiền sau"] = parts[-1]      # 1.680.480
                         data["Tiền thuế"] = parts[-2]        # 124.480
                         data["Số tiền trước Thuế"] = parts[0] # 1.656.000 (NOT parts[-3])
                    elif len(parts) >= 3 and all(c in '0123456789.,' for c in ''.join(parts)):
                         # 3-column: before TAX total
                         data["Số tiền sau"] = parts[-1]
                         data["Tiền thuế"] = parts[-2]
                         data["Số tiền trước Thuế"] = parts[0]
                    else:
                         data["Số tiền sau"] = val
                break
        
        # HDDT fix: "Tiền thanh toán:" multi-column format (e.g. "Tiền thanh toán: 0 0 0 0 200.000 0 200.000")
        # Take the last non-zero value from the row
        if not data["Số tiền sau"]:
            tien_tt_match = re.search(r'Tiền thanh toán[:\s]+([\d\s.,]+)', full_text, re.IGNORECASE)
            if tien_tt_match:
                nums_str = tien_tt_match.group(1).strip()
                nums_list = re.findall(r'[\d]+(?:[.,][\d]+)*', nums_str)
                for n in reversed(nums_list):
                    if parse_money(n) > 0:
                        data["Số tiền sau"] = n
                        break
        
        # SPECIAL CASE: Hộ Kinh Doanh with Tax Reduction Note (Nghị quyết 204/2025/QH15)
        # e.g. "Cộng tiền bán hàng hóa, dịch vụ: 2.289.962" -> This is the final amount to pay
        if not data["Số tiền sau"] or not data["Số tiền trước Thuế"]:
             # Check for "Cộng tiền bán hàng hóa, dịch vụ" which is common in direct sales invoices
             direct_sales_match = re.search(r'Cộng tiền bán hàng hóa, dịch vụ[:\s]*([\d\.,]+)', full_text, re.IGNORECASE)
             if direct_sales_match:
                 amount = direct_sales_match.group(1)
                 # If we haven't set "Số tiền sau", use this. 
                 # Usually for Hộ Kinh Doanh, total payment = total goods amount (minus discount if any, but usually final)
                 if not data["Số tiền sau"]:
                     data["Số tiền sau"] = amount
                 if not data["Số tiền trước Thuế"]:
                     data["Số tiền trước Thuế"] = amount
                 # If extracted "Tiền thuế" is empty, it might be 0 or calculated from reduction note, 
                 # but usually direct sales don't list VAT separately like deductive invoices. 
                 # We leave VAT empty or 0 if not found.

        # SPECIAL PATTERN: Before Tax + VAT on one line (File 1226-TK-200k.pdf)
        # Cộng tiền hàng hóa, dịch vụ: 219.907 17.593
        # Prioritize this summary line as it matches User's preferred values (rounded)
        double_match = re.search(r'Cộng tiền hàng hóa, dịch vụ[:\s]*([\d\.,]+)\s+([\d\.,]+)', full_text, re.IGNORECASE)
        if double_match:
             # Check if the second number looks like money (digits/dots)
             # Force overwrite to ensure we get the summary values
             data["Số tiền trước Thuế"] = double_match.group(1)
             data["Tiền thuế"] = double_match.group(2)
        
        # For SALES INVOICE (no VAT): if Số tiền sau is empty but we have before tax amount
        if not data["Số tiền sau"] and data["Số tiền trước Thuế"]:
            if "SALES INVOICE" in full_text or "HÓA ĐƠN BÁN HÀNG" in full_text:
                data["Số tiền sau"] = data["Số tiền trước Thuế"]
            else:
                sales_match = re.search(r'Cộng tiền bán hàng[^:]*[:\s]*([\d\.,]+)', full_text)
                if sales_match:
                    data["Số tiền sau"] = sales_match.group(1)
        
        # Calculate and validate money values
        # Validate money values - must be >= 1000
        for col in ["Số tiền trước Thuế", "Tiền thuế", "Số tiền sau"]:
            val = parse_money(data[col])
            if val is not None and val < 1000:
                data[col] = ""  # Invalid, clear it
        
        # Calculate Số tiền sau if not found but we have before tax and VAT
        if not data["Số tiền sau"]:
            before = parse_money(data["Số tiền trước Thuế"])
            vat = parse_money(data["Tiền thuế"])
            if before is not None and vat is not None:
                data["Số tiền sau"] = format_money(before + vat)
            elif before is not None and vat is None:
                # No VAT found yet. BUT check if we have "Thuế khác" indicating a rate!
                # If we have a rate (e.g. "10"), we should NOT assume Total = PreTax yet.
                if data["Thuế khác"] and data["Thuế khác"].strip() in ["10", "5", "8"]:
                     pass # Wait for calculation
                
                # If Thuế khác is same as Total or Tax, it's noise
                if data["Thuế khác"]:
                     val_num = parse_money(data["Thuế khác"])
                     total_num = parse_money(data["Số tiền sau"])
                     tax_num = parse_money(data["Tiền thuế"])
                     if val_num and (val_num == total_num or val_num == tax_num):
                          data["Thuế khác"] = ""
                else:
                    # No VAT, total = before tax
                    data["Số tiền sau"] = data["Số tiền trước Thuế"]
        
        # REVERSE CASE: If we have Số tiền sau (total) but no Số tiền trước Thuế (before tax)
        # and no VAT was found, then this is a non-VAT invoice, so set pre-tax = post-tax
        if data["Số tiền sau"] and not data["Số tiền trước Thuế"]:
            after = parse_money(data["Số tiền sau"])
            vat = parse_money(data["Tiền thuế"])
            if after is not None and (vat is None or vat == 0):
                # No VAT found or VAT is 0, so pre-tax = post-tax
                data["Số tiền trước Thuế"] = data["Số tiền sau"]
            elif after is not None and vat is not None:
                # VAT exists, so calculate pre-tax = post-tax - VAT
                # VAT exists, so calculate pre-tax = post-tax - VAT
                data["Số tiền trước Thuế"] = format_money(after - vat)
        
        # SPECIAL FIX for PSD.pdf where total is hidden in garbage
        # We recovered "2,950,000" from garbage but regex didn't catch it as Total.
        # Check if we have a valid recovered amount in line items but no Invoice Total?
        # Actually, let's look for the specific garbage string containing the Total
        if not data["Số tiền sau"]:
             garbage_total = re.search(r"0'}([\d\.,]+)'\}'\}", full_text)
             if garbage_total:
                 val = garbage_total.group(1)
                 data["Số tiền sau"] = val
                 # Use this as PreTax too if missing (or calc tax)
                 if not data["Số tiền trước Thuế"]:
                      data["Số tiền trước Thuế"] = val

        # FALLBACK: If "Số tiền trước Thuế" or "Số tiền sau" is still missing, 
        # try to sum up the Line Items!
        if (not data["Số tiền trước Thuế"] or not data["Số tiền sau"]) and line_items:
            print("  -> Calculating totals from line items...")
            total_items = 0
            for item in line_items:
                amt = parse_money(item.get("amount", "0"))
                if amt:
                    total_items += amt
            
            if total_items > 0:
                if not data["Số tiền trước Thuế"] and not data["Số tiền sau"]:
                     # Assume line items are pre-tax (standard) or post-tax? 
                     # Usually line items amount column is Before Tax.
                     data["Số tiền trước Thuế"] = format_money(total_items)
                elif not data["Số tiền trước Thuế"]:
                     data["Số tiền trước Thuế"] = format_money(total_items)
                elif not data["Số tiền sau"]:
                     # If we have PreTax but no PostTax, we need Tax to calc Total.
                     # If we just calculated PreTax, let's see if we can calc Total
                     pass
                     
        # Re-run Tax Calculation in case we just populated Pre-Tax from items
        if data["Tiền thuế"] and not data["Số tiền sau"] and data["Số tiền trước Thuế"]:
             b = parse_money(data["Số tiền trước Thuế"])
             t = parse_money(data["Tiền thuế"])
             if b and t:
                 data["Số tiền sau"] = format_money(b + t)
        
        # --- NEW STRATEGY: PARSE SUMMARY TABLES (Footer) ---
        # Many invoices (like PSD.pdf and 0318...pdf) have a summary block with tax rates
        # Pattern: "Hàng hóa ... 8% ... [PreTax] ... [Tax] ... [Total]"
        # Pattern: "Cộng HHDV ... 10% ... [PreTax] ... [Tax]"
        
        # 1. Parse Detail Lines for Tax Rates (8%, 10%, 5%, 0%)
        # Look for lines containing "8%" or "10%" followed by multiple money numbers
        summary_lines = re.findall(r'(?:Hàng hóa|Cộng HHDV|Thuế suất|Total amount).*?(10%|8%|5%|0%).*?([\d\.,]+)\s+([\d\.,]+)(?:\s+([\d\.,]+))?', full_text, re.IGNORECASE)
        
        tax_total_calc = 0
        pre_tax_total_calc = 0
        
        for rate_str, num1, num2, num3 in summary_lines:
            # Usually: Rate, PreTax, Tax, [Total] OR Rate, [Total], [Tax]
            # Heuristic: Tax is usually smaller than PreTax. 
            # num1, num2, num3 are strings.
            try:
                vals = [parse_money(n) for n in [num1, num2, num3] if n]
                vals.sort() # Sorted: [Smallest, Medium, Largest]
                
                # Smallest is likely Tax (if > 0)
                # Largest is Total (or PreTax if Total missing)
                
                # If we have 2 numbers: PreTax and Tax
                # If we have 3 numbers: PreTax, Tax, Total
                
                if len(vals) >= 2:
                    current_tax = vals[0]
                    current_pre = vals[-1] # Largest is PreTax (if 2 nums) or Total (if 3 nums)? 
                    # Actually, if 3 nums: Tax, PreTax, Total. PreTax is middle.
                    if len(vals) == 3:
                         current_pre = vals[1]
                    
                    # Store in specific tax column
                    rate_key = f"Thuế {rate_str}"
                    data[rate_key] = format_money(current_tax)
                    
                    tax_total_calc += current_tax
                    pre_tax_total_calc += current_pre
                    
                    # print(f"  -> Found Summary Line: {rate_str} | Tax: {current_tax} | Pre: {current_pre}")
            except:
                pass
        
        # If we found summary data, assume it's the source of truth for Totals
        if tax_total_calc > 0:
            if not data["Tiền thuế"] or parse_money(data["Tiền thuế"]) != tax_total_calc:
                 data["Tiền thuế"] = format_money(tax_total_calc)
        
        # 2. Parse Grand Total Line with multiple numbers
        # Pattern: "Tổng cộng tiền ... [PreTax] [Tax] [Total]" (common in 0318...pdf)
        grand_total_match = re.search(r'(?:Tổng cộng tiền|Grand total).*?([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)', full_text, re.IGNORECASE)
        if grand_total_match:
             v1 = parse_money(grand_total_match.group(1))
             v2 = parse_money(grand_total_match.group(2))
             v3 = parse_money(grand_total_match.group(3))
             
             vals = [v for v in [v1, v2, v3] if v is not None]
             vals.sort()
             if len(vals) == 3:
                 # Tax, PreTax, Total
                 data["Tiền thuế"] = format_money(vals[0])
                 data["Số tiền trước Thuế"] = format_money(vals[1])
                 data["Số tiền sau"] = format_money(vals[2])
                 print(f"  -> Found Grand Total Line: Total={vals[2]}, Tax={vals[0]}")

        # Missing Lookup Code for PSD.pdf (e5100...)
        if not data["Mã tra cứu"]:
             # Pattern: "nhập mã ...", "key in the provided code ... : [code]"
             code_match = re.search(r'(?:nhập mã|provided code).*?([a-f0-9]{30,})', full_text, re.IGNORECASE)
             if code_match:
                 data["Mã tra cứu"] = code_match.group(1)
             
             # Additional Lookup Code Pattern (PC-...)
             # Example: PC-260107070845-3863477
             pc_match = re.search(r'Mã tra cứu[:\s]*([A-Z0-9-]+)', full_text, re.IGNORECASE)
             if pc_match:
                 data["Mã tra cứu"] = pc_match.group(1)
        
        # FINAL: If Tax Rate found (e.g. "Thuế khác": "10") but Column Empty, fill it
        # This fixes 0318...pdf where "Thuế khác" picked up "10" but didn't fill "Thuế 10%"
        if data["Thuế khác"] in ["10", "8", "5", "0"]:
            rate_key = f"Thuế {data['Thuế khác']}%"
            if not data[rate_key] and data["Tiền thuế"]:
                data[rate_key] = data["Tiền thuế"]
                data["Thuế khác"] = ""
            elif not data[rate_key] and data["Số tiền trước Thuế"]:
                # Calculate tax from rate
                try:
                    rate = int(data["Thuế khác"])
                    pre = parse_money(data["Số tiền trước Thuế"])
                    if pre:
                        calc_tax = int(round(pre * rate / 100))
                        data[rate_key] = format_money(calc_tax)
                        if not data["Tiền thuế"]:
                             data["Tiền thuế"] = format_money(calc_tax)
                        
                        # Use loose check for Total assignment
                        current_total = parse_money(data["Số tiền sau"])
                        pre_val = parse_money(data["Số tiền trước Thuế"])
                        
                        # If Total is empty OR Total == PreTax (from premature assignment), update it!
                        if not data["Số tiền sau"] or (current_total and pre_val and abs(current_total - pre_val) < 100):
                             data["Số tiền sau"] = format_money(pre + calc_tax)
                        
                        data["Thuế khác"] = ""
                except:
                    pass

        
        
        # Store line items for multi-row expansion
        if services:
            line_items = services
            # POST-PROCESS: Clean garbage from items
            for item in line_items:
                for k in ["name", "amount"]:
                    val = item.get(k, "")
                    if isinstance(val, str) and ("0'}" in val or "}'}" in val):
                        # Standard Garbage Removal using Regex
                        garbage_match = re.search(r"0'}.*?([\d\.,]+).*?'\}'\}", val, re.DOTALL)
                        if garbage_match:
                             # We found hidden numbers in garbage, but we trust the Footer Summary Table now for Totals.
                             # So just clean the item value.
                             item[k] = val.replace(garbage_match.group(0), "").strip()
                        else:
                             item[k] = val.replace("0'}", "").replace("'}'}", "").replace("'}", "").replace("{'", "").strip()
            
            # Aggregate taxes from line items if detected (e.g. for invoices with no summary table)
            tax_map = {0: 0, 5: 0, 8: 0, 10: 0}
            has_item_tax = False
            
            for item in line_items:
                r_str = item.get('tax_rate')
                amt_str = item.get('amount')
                # print(f"  [DEBUG LOOP] Item Rate: '{r_str}', Amt: '{amt_str}'")
                if r_str and amt_str:
                    try:
                        amt = parse_money(amt_str)
                        r = int(r_str)
                        # print(f"    [DEBUG] Item Amount: {amt}, Rate: {r}")
                        if r in tax_map:
                             # Calculate Tax = Amount * Rate / 100
                             # Note: This is an estimation. Ideally we parse the tax amount column too but that varies wildly in format.
                             tax_val = int(round(amt * r / 100))
                             tax_map[r] += tax_val
                             has_item_tax = True
                    except Exception as e:
                        # print(f"    [DEBUG ERROR] Item process failed: {e}")
                        pass
            
            if has_item_tax:
                # Fill missing tax buckets
                for r in [0, 5, 8, 10]:
                    key = f"Thuế {r}%"
                    if tax_map[r] > 0:
                        # Only overwrite if empty or significantly different (likely better data from items than bad footer parse)
                        curr_val = parse_money(data[key])
                        diff = abs(curr_val - tax_map[r])
                        
                        # Overwrite strategy:
                        # 1. If Curr is 0/Empty -> Overwrite
                        # 2. If Diff is Huge (> 50% of Calc) -> Trust Calc (Fix Garbage regex capture)
                        # NOTE: Do NOT overwrite small diffs. Trust the OCR/Document if it's close.
                        if curr_val == 0 or diff > tax_map[r] * 0.5:
                             data[key] = format_money(tax_map[r])
                             
                # Recalculate Total Tax if it looks wrong or empty
                total_item_tax = sum(tax_map.values())
                curr_total_tax = parse_money(data["Tiền thuế"])
                
                # If total tax is missing or significantly smaller than item sum (e.g. captured only one rate), update it
                if curr_total_tax == 0 or (total_item_tax > curr_total_tax and total_item_tax > 1000):
                     data["Tiền thuế"] = format_money(total_item_tax)
                     print(f"  [AUTO-AGGR] Aggregated Tax from items: {total_item_tax}")

            # FINAL CHECK: Sanity check Tax Amount (Run AFTER post-process updates)
            if data["Tiền thuế"] and (data["Số tiền trước Thuế"] or data["Số tiền sau"]):
                 try:
                    t_val = parse_money(data["Tiền thuế"])
                    # Use PreTax, or infer from Total if Tax is huge
                    p_val = parse_money(data["Số tiền trước Thuế"]) 
                    if not p_val and data["Số tiền sau"]:
                         # Assume Total > Tax
                         p_val = parse_money(data["Số tiền sau"])
                    
                    if t_val and p_val and p_val > 10000 and t_val >= p_val: # Strict Check: Tax >= PreTax/Total
                        print(f"  -> Discarding suspicious Tax Amount: {data['Tiền thuế']} (Validation Failed: > Amount)")
                        data["Tiền thuế"] = ""
                        for c in ["Thuế 0%", "Thuế 5%", "Thuế 8%", "Thuế 10%", "Thuế khác"]:
                             if parse_money(data[c]) == t_val:
                                 data[c] = ""
                 except:
                    pass

    except Exception as e:
        print(f"Error processing {filename}: {e}")
    
    # RE-CALCULATE TOTAL TAX from Components if missing
    if not data["Tiền thuế"]:
         calc_tax = 0
         for c in ["Thuế 0%", "Thuế 5%", "Thuế 8%", "Thuế 10%", "Thuế khác"]:
             calc_tax += parse_money(data.get(c))
         
         if calc_tax > 0:
             data["Tiền thuế"] = format_money(calc_tax)
             print(f"  [AUTO-AGGR] Inferred Total Tax from breakdown: {calc_tax}")

    # RE-CALCULATE TAX RATE if missing (Final Pass)
    # This runs after all other fallbacks/aggregations to catch cases where Pre/Tax were inferred but Rate wasn't set.
    if data["Tiền thuế"] and data["Số tiền trước Thuế"] and not any(data.get(c) for c in ["Thuế 0%", "Thuế 5%", "Thuế 8%", "Thuế 10%"]):
        try:
             t_val = parse_money(data["Tiền thuế"])
             p_val = parse_money(data["Số tiền trước Thuế"])
             if t_val > 0 and p_val > 0:
                 rate = round(t_val / p_val * 100)
                 # Allow slight tolerance if needed, but rounding usually handles it
                 if rate in [0, 5, 8, 10]:
                      key = f"Thuế {rate}%"
                      data[key] = data["Tiền thuế"]
        except Exception:
             pass

    # FINAL CLEANUP: Clean all data fields
    for k, v in data.items():
        if isinstance(v, str):
            data[k] = clean_string_value(v)
            
    # Extra cleanup for Tax Rate fields
    if data.get("Thuế khác"):
         # If it's just a rate like "10" and we already have "Thuế 10%" filled, clear "Thuế khác"
         if data["Thuế khác"] in ["10", "8", "5", "0"]:
              rate_key = f"Thuế {data['Thuế khác']}%"
              if data.get(rate_key):
                   data["Thuế khác"] = ""
         
         # If it matches Total or Tax, it's noise
         v_num = parse_money(data["Thuế khác"])
         t_num = parse_money(data.get("Số tiền sau"))
         tax_num = parse_money(data.get("Tiền thuế"))
         if v_num and (v_num == t_num or v_num == tax_num):
              data["Thuế khác"] = ""

    # FINAL MONEY CLEANUP: Ensure all money fields are properly formatted and decimal suffixes removed
    money_keys = ["Số tiền trước Thuế", "Tiền thuế", "Thuế 0%", "Thuế 5%", "Thuế 8%", "Thuế 10%", "Thuế khác", "Số tiền sau"]
    for k in money_keys:
        if k in data and data[k]:
             orig = data[k]
             # print(f"[DEBUG CLEANUP] Key: {k}, Orig: '{orig}'")
             val = parse_money(data[k])
             # print(f"    -> Parsed: {val}")
             
             if val != 0:
                 data[k] = format_money(val)
                 # print(f"    -> Formatted: '{data[k]}'")
             elif val == 0 and data[k] not in ["0", "0.0", "0,0"]:
                 pass

    # FINAL CLEANUP for Seller Name (Global)
    # 0. Pre-clean: If current seller is just a blacklisted string, clear it so Rescue can work
    if data.get("Đơn vị bán"):
         s = data["Đơn vị bán"]
         if any(bad in s.lower() for bad in ['đã được ký điện tử bởi', 'được ký bởi', 'ký bởi công ty', 'digitally signed by']):
             data["Đơn vị bán"] = ""

    # RESCUE: If Seller is empty, check for known short names (Quán, Shop)
    if not data["Đơn vị bán"]:
         # "Quán 87", "Quán Cơm", "Shop ABC"
         if "Quán 87" in full_text:
             data["Đơn vị bán"] = "Quán 87"

    if data.get("Đơn vị bán"):
        s = data["Đơn vị bán"]
        # Remove (Issued) : prefixes
        s = re.sub(r'^\s*\(?Issued\)?\s*[:\.\-]\s*', '', s, flags=re.IGNORECASE).strip()
        # Remove "Đơn vị bán" prefix if leaked
        s = re.sub(r'^\s*[\(\[]?\s*(?:Seller|Company|Người bán|Doanh nghiệp|Tên đơn vị|Đơn vị bán)\s*[\)\]]?\s*[:\.\-]?\s*', '', s, flags=re.IGNORECASE)

        # Remove Ký hiệu suffix if present (e.g. "... Ký hiệu: 1K25TAN")
        if "Ký hiệu:" in s:
            s = s.split("Ký hiệu:")[0].strip()
        
        # Check blacklist - if blacklisted, clear it
        if any(bad in s.lower() for bad in ['đã được ký điện tử bởi', 'được ký bởi', 'ký bởi công ty', 'digitally signed by']):
             s = "" 
        
        data["Đơn vị bán"] = s

    return data, line_items


def validate_invoice_data(data):
    """
    Validate extracted invoice data and return list of issues.
    Each issue is a dict: {"field": str, "severity": "error"|"warning", "message": str}
    """
    issues = []
    
    def _parse_money(s):
        if not s: return 0
        s = str(s).strip().replace(',', '').replace('.', '')
        try: return int(s)
        except: return 0
    
    # === ERRORS (missing critical fields) ===
    seller = data.get("Đơn vị bán", "").strip()
    if not seller or len(seller) < 5:
        issues.append({"field": "Đơn vị bán", "severity": "error", "message": "Thiếu hoặc quá ngắn"})
    
    if not data.get("Số hóa đơn", "").strip():
        issues.append({"field": "Số hóa đơn", "severity": "error", "message": "Thiếu số hóa đơn"})
    
    date_val = data.get("Ngày hóa đơn", "").strip()
    if not date_val:
        issues.append({"field": "Ngày hóa đơn", "severity": "error", "message": "Thiếu ngày hóa đơn"})
    elif not re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', date_val):
        issues.append({"field": "Ngày hóa đơn", "severity": "warning", "message": f"Định dạng ngày không chuẩn: {date_val}"})
    
    before_tax = _parse_money(data.get("Số tiền trước Thuế", ""))
    if before_tax == 0:
        issues.append({"field": "Số tiền trước Thuế", "severity": "error", "message": "Số tiền trước thuế = 0 hoặc trống"})
    
    after_tax = _parse_money(data.get("Số tiền sau", ""))
    if after_tax == 0:
        issues.append({"field": "Số tiền sau", "severity": "error", "message": "Số tiền sau thuế = 0 hoặc trống"})
    
    # === WARNINGS (suspicious but not critical) ===
    mst = data.get("Mã số thuế", "").strip()
    if not mst:
        issues.append({"field": "Mã số thuế", "severity": "warning", "message": "Thiếu mã số thuế"})
    elif not re.match(r'^[\d\-]{10,14}$', mst):
        issues.append({"field": "Mã số thuế", "severity": "warning", "message": f"MST không hợp lệ ({len(mst)} ký tự)"})
    
    if not data.get("Link lấy hóa đơn", "").strip():
        issues.append({"field": "Link lấy hóa đơn", "severity": "warning", "message": "Thiếu link tra cứu"})
    
    if not data.get("Mã tra cứu", "").strip():
        issues.append({"field": "Mã tra cứu", "severity": "warning", "message": "Thiếu mã tra cứu"})
    
    if not data.get("Ký hiệu", "").strip():
        issues.append({"field": "Ký hiệu", "severity": "warning", "message": "Thiếu ký hiệu"})
    
    # Amount logic check
    if before_tax > 0 and after_tax > 0 and after_tax < before_tax:
        issues.append({"field": "Số tiền sau", "severity": "warning", 
                       "message": f"Số tiền sau ({after_tax:,}) < trước thuế ({before_tax:,})"})
    
    return issues


def clean_string_value(val):
    """Clean string values from control characters and excessive whitespace."""
    if not isinstance(val, str):
        return val
    # Remove control characters like \r, \xad
    val = val.replace('\r', '').replace('\xad', '').replace('\t', ' ')
    # Normalize whitespace
    val = " ".join(val.split())
    return val



def format_excel_output(file_path):
    """Format the Excel output file with professional styles and merge cells."""
    print(f"Applying professional formatting to {file_path}...")
    try:
        wb = openpyxl.load_workbook(file_path)
        if "Hóa đơn" in wb.sheetnames:
            ws = wb["Hóa đơn"]
        else:
            ws = wb.active

        # Define Styles
        header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
        header_fill = PatternFill("solid", fgColor="4F81BD")  # Professional Blue
        border_style = Side(style='thin', color="000000")
        border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        
        # Column widths for layout with VAT breakdown:
        # A=Tên file, B=Ngày, C=Số HĐ, D=Đơn vị bán, E=Phân loại
        # F=Trước thuế, G=Thuế 0%, H=Thuế 5%, I=Thuế 8%, J=Thuế 10%, K=Thuế khác
        # L=Tiền thuế, M=Sau thuế, N=Link, O=Mã tra cứu, P=MST, Q=Mã CQT, R=Ký hiệu
        widths = {
            'A': 30, 'B': 12, 'C': 15, 'D': 40, 'E': 18,
            'F': 18, 'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12,
            'L': 12, 'M': 15, 'N': 18, 'O': 15, 'P': 20, 'Q': 15, 'R': 15, 'S': 12
        }
        
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width

        # Format Header Row (row 1)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Add Filter
        ws.auto_filter.ref = ws.dimensions
        
        # No need for merge logic since each invoice is now one row
        
        # Format Data Rows - borders and number formatting
        money_cols_idx = [6, 7, 8, 9, 10, 11, 12, 13]  # F through M (all money columns)
        center_cols_idx = [2, 3, 5, 16, 18]  # B, C, E, P, R
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                # Skip merged cells (they don't have col_idx)
                if not hasattr(cell, 'col_idx'):
                    continue
                    
                cell.border = border
                cell.font = Font(name="Arial", size=10)
                
                # Number format for money columns
                if cell.col_idx in money_cols_idx:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif cell.col_idx in center_cols_idx:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

        wb.save(file_path)
        print(" -> Formatting complete with merged cells.")
        
    except Exception as e:
        print(f"Error formatting Excel: {e}")
        import traceback
        traceback.print_exc()

def main():
    # Fix Windows console encoding for Vietnamese characters
    import sys
    sys.stdout.reconfigure(encoding='utf-8')
    
    # Input folder for PDF files - users should place new invoices here
    input_folder = r"D:\hoadon\invoices_input"
    
    # Output folder for exported Excel
    output_folder = r"D:\hoadon"
    
    # Create input folder if not exists
    if not os.path.exists(input_folder):
        os.makedirs(input_folder)
        print(f"Created input folder: {input_folder}")
        print("Please add PDF invoice files to this folder and run again.")
        return
    
    # Get all PDF files from input folder
    pdf_files = [f for f in os.listdir(input_folder) if f.lower().endswith('.pdf')]
    
    if not pdf_files:
        print(f"No PDF files found in: {input_folder}")
        print("Please add PDF invoice files to this folder and run again.")
        return
        
    print(f"Processing {len(pdf_files)} PDF files from: {input_folder}\\n")
    
    all_rows = []  # Will contain expanded rows (one per line item)
    
    for pdf_file in pdf_files:
        pdf_path = os.path.join(input_folder, pdf_file)
        print(f"Processing: {pdf_file}")
        
        data, line_items = extract_invoice_data(pdf_path)
        
        # Classify invoice based on line items
        if line_items:
            all_item_names = " ".join([item.get("name", "") for item in line_items])
            data["Phân loại"] = classify_content(all_item_names, data.get("Đơn vị bán", ""))
        else:
            data["Phân loại"] = "Khác"
        
        all_rows.append(data)
        
        # Show status
        item_count = len(line_items) if line_items else 0
        seller_display = data['Đơn vị bán'][:30] if data['Đơn vị bán'] and len(data['Đơn vị bán']) > 0 else 'N/A'
        pv_display = f", PV: {data['Phí PV']}" if data.get('Phí PV') else ""
        print(f"  -> Ngay: {data['Ngày hóa đơn']}, So: {data['Số hóa đơn']}, Category: {data['Phân loại']}, DonViBan: {seller_display}{pv_display}...")
    
    # Create DataFrame
    df = pd.DataFrame(all_rows)
    
    # Reorder columns
    columns = [
        "Tên file", "Ngày hóa đơn", "Số hóa đơn", "Đơn vị bán", "Phân loại",
        "Số tiền trước Thuế", "Thuế 0%", "Thuế 5%", "Thuế 8%", "Thuế 10%", "Thuế khác",
        "Tiền thuế", "Số tiền sau", "Link lấy hóa đơn",
        "Mã tra cứu", "Mã số thuế", "Mã CQT", "Ký hiệu"
    ]
    df = df[columns]
    
    # Format money columns - convert to number and add comma separators
    money_columns = ["Số tiền trước Thuế", "Thuế 0%", "Thuế 5%", "Thuế 8%", "Thuế 10%", "Thuế khác", "Tiền thuế", "Số tiền sau"]
    for col in money_columns:
        def convert_to_number(x):
            if pd.isna(x) or x == '':
                return None
            x_str = str(x).strip()
            # Detect format: 
            # - Vietnamese decimal: 17.592,59 (comma before 2 digits at end)
            # - Thousands only: 79,600 (comma before 3 digits at end)
            import re
            # If comma followed by exactly 2 digits at end, it's decimal
            if re.search(r',\d{2}$', x_str):
                # Vietnamese format: . = thousands, , = decimal
                x_str = x_str.replace('.', '').replace(',', '.')
            else:
                # Comma is thousands separator, just remove both . and ,
                x_str = x_str.replace('.', '').replace(',', '')
            try:
                return round(float(x_str))
            except (ValueError, TypeError):
                return x
        df[col] = df[col].apply(convert_to_number)
    
    # Export to Excel
    output_file = os.path.join(output_folder, "hoadon_tonghop.xlsx")
    try:
        df.to_excel(output_file, index=False, sheet_name="Hóa đơn")
    except PermissionError:
        print(f"\nWARNING: Could not save to '{output_file}' because it is open.")
        output_file = os.path.join(output_folder, "hoadon_tonghop_new.xlsx")
        print(f"Saving to '{output_file}' instead.")
        df.to_excel(output_file, index=False, sheet_name="Hóa đơn")
    
    # Print summary
    print(f"\n{'='*50}")
    print(f"SUMMARY:")
    unique_files = df["Tên file"].nunique()
    print(f"  Total rows: {len(df)} (from {unique_files} invoices)")
    for col in ["Ngày hóa đơn", "Số hóa đơn", "Đơn vị bán", "Phân loại", "Số tiền sau"]:
        empty_count = (df[col] == '').sum() + df[col].isna().sum()
        pct = (1 - empty_count/len(df)) * 100
        print(f"  {col}: {pct:.0f}% filled ({len(df)-empty_count}/{len(df)})")
    print(f"\nExported to: {output_file}")
    
    # Apply professional formatting
    format_excel_output(output_file)

if __name__ == "__main__":
    main()
