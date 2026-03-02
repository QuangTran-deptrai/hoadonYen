import streamlit as st
import pandas as pd
import io
import os
import logging
import sys
import re
from extract_invoices import extract_invoice_data, classify_content, validate_invoice_data
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Configure logging to stdout
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

# Configure page
st.set_page_config(page_title="Invoice Extractor", page_icon="🧾", layout="wide")

# OCR diagnostics in sidebar
with st.sidebar:
    st.markdown("### 🔧 OCR Status")
    try:
        from extract_invoices import OCR_AVAILABLE, POPPLER_PATH
        if OCR_AVAILABLE:
            st.success("✅ OCR Available")
            import pytesseract
            import subprocess
            # Check Tesseract version
            try:
                tess_ver = pytesseract.get_tesseract_version()
                st.info(f"Tesseract: v{tess_ver}")
            except Exception as e:
                st.warning(f"Tesseract: {e}")
            # Check Vietnamese language
            try:
                langs = pytesseract.get_languages()
                has_vie = 'vie' in langs
                st.info(f"Languages: {', '.join(langs[:5])}")
                if has_vie:
                    st.success("✅ Vietnamese OK")
                else:
                    st.warning("⚠️ Vietnamese not found")
            except Exception:
                st.warning("⚠️ Cannot check languages")
            # Check Poppler
            if POPPLER_PATH:
                st.info(f"Poppler: {POPPLER_PATH}")
            else:
                try:
                    result = subprocess.run(['pdftoppm', '-v'], capture_output=True, text=True, timeout=5)
                    st.success("✅ Poppler (system)")
                except Exception:
                    st.error("❌ Poppler not found")
        else:
            st.error("❌ OCR not available")
    except Exception as e:
        st.error(f"❌ Error: {e}")

# Category options for dropdown
CATEGORY_OPTIONS = [
    "Tự động nhận diện",  # Auto-detect based on invoice content
    "Dịch vụ ăn uống",
    "Dịch vụ phòng nghỉ", 
    "Hoa tươi",
    "Thẻ cào điện thoại",
    "Xăng xe",
    "Quà tặng",
    "Khác (Nhập tay)"
]

# Initialize Session State
if "processing_complete" not in st.session_state:
    st.session_state["processing_complete"] = False
if "processed_df" not in st.session_state:
    st.session_state["processed_df"] = None
if "report_type" not in st.session_state:
    st.session_state["report_type"] = "Kế toán"
if "validation_results" not in st.session_state:
    st.session_state["validation_results"] = []

# --- Main Application Logic (no login required) ---

# Sidebar
with st.sidebar:
    st.markdown("**Invoice Extractor**")
    st.markdown("---")
    st.caption("Phan tich va trich xuat du lieu tu hoa don PDF")

# App Title
st.title("Invoice Extraction Tool")

# --- WIZARD FLOW ---

if st.session_state["processing_complete"] and st.session_state["processed_df"] is not None:
    # === STEP 4: RESULTS & EXPORT ===
    st.markdown("### ✅ Kết quả xử lý")
    
    col_res1, col_res2 = st.columns([1, 4])
    with col_res1:
        if st.button("⬅️ Làm việc với file khác"):
            st.session_state["processing_complete"] = False
            st.session_state["processed_df"] = None
            st.session_state["validation_results"] = []
            st.rerun()
    
    df = st.session_state["processed_df"]
    validations = st.session_state.get("validation_results", [])
    
    # === VALIDATION SUMMARY ===
    if validations:
        # Count errors and warnings across all files
        total_errors = 0
        total_warnings = 0
        files_with_issues = {}
        
        for v in validations:
            fname = v["file"]
            issues = v["issues"]
            if issues:
                errs = sum(1 for i in issues if i["severity"] == "error")
                warns = sum(1 for i in issues if i["severity"] == "warning")
                total_errors += errs
                total_warnings += warns
                if errs > 0 or warns > 0:
                    files_with_issues[fname] = {"errors": errs, "warnings": warns, "details": issues}
        
        if total_errors > 0 or total_warnings > 0:
            st.markdown("### 🔍 Phát hiện lỗi / cảnh báo")
            
            # Summary metrics
            col_m1, col_m2, col_m3 = st.columns(3)
            with col_m1:
                st.metric("🔴 Lỗi", total_errors)
            with col_m2:
                st.metric("🟡 Cảnh báo", total_warnings)
            with col_m3:
                st.metric("📄 File có vấn đề", len(files_with_issues))
            
            # Expandable details per file
            for fname, info in files_with_issues.items():
                error_count = info["errors"]
                warn_count = info["warnings"]
                badge = ""
                if error_count > 0:
                    badge += f"🔴 {error_count} lỗi "
                if warn_count > 0:
                    badge += f"🟡 {warn_count} cảnh báo"
                
                with st.expander(f"📄 **{fname}** — {badge}", expanded=(error_count > 0)):
                    for issue in info["details"]:
                        icon = "🔴" if issue["severity"] == "error" else "🟡"
                        st.markdown(f"{icon} **{issue['field']}**: {issue['message']}")
            
            st.info("💡 **Bạn có thể sửa trực tiếp trong bảng bên dưới** trước khi xuất Excel.")
            st.divider()
        else:
            st.success("✅ Tất cả hóa đơn đều hợp lệ! Không phát hiện lỗi.")
            st.divider()
    
    # === EDITABLE DATA TABLE ===
    st.markdown("### ✏️ Kiểm tra & Chỉnh sửa dữ liệu")
    st.caption("Nhấp đúp vào ô bất kỳ để sửa. Dữ liệu sẽ được cập nhật ngay khi bạn nhấn Enter hoặc click ra ngoài.")
    
    # Build "Trạng thái" column based on validation
    status_map = {}
    if validations:
        for v in validations:
            fname = v["file"]
            issues = v["issues"]
            if any(i["severity"] == "error" for i in issues):
                status_map[fname] = "🔴 Lỗi"
            elif any(i["severity"] == "warning" for i in issues):
                status_map[fname] = "🟡 Cảnh báo"
            else:
                status_map[fname] = "✅ OK"
    
    # Find the filename column
    fname_col = "Tên file" if "Tên file" in df.columns else None
    
    # Add status column if we have validations
    display_df = df.copy()
    if fname_col and status_map:
        display_df.insert(0, "Trạng thái", display_df[fname_col].map(status_map).fillna("✅ OK"))
    
    # Use st.data_editor for inline editing
    edited_df = st.data_editor(
        display_df,
        width='stretch',
        num_rows="fixed",
        key="invoice_editor",
        height=min(800, 45 + len(display_df) * 35),
    )
    
    # Remove status column from export data (it's display-only)
    if "Trạng thái" in edited_df.columns:
        export_df = edited_df.drop(columns=["Trạng thái"])
    else:
        export_df = edited_df
    
    st.divider()
    
    # === EXCEL EXPORT (uses edited data) ===
    report_type = st.session_state.get("report_type", "Kế toán")
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        export_df.to_excel(writer, index=False, sheet_name="Hóa đơn")
        worksheet = writer.sheets["Hóa đơn"]
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        border_style = Side(style='thin', color="000000")
        border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        
        # Format Header
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            
        worksheet.freeze_panes = 'A2'
        worksheet.auto_filter.ref = worksheet.dimensions

        if report_type == "Kinh doanh":
            # === BUSINESS FORMAT ===
            widths = {'A': 15, 'B': 20, 'C': 30, 'D': 12, 'E': 10, 'F': 25, 'G': 15, 
                      'H': 15, 'I': 12, 'J': 12, 'K': 12, 'L': 12, 'M': 10, 'N': 15, 
                      'O': 15, 'P': 30, 'Q': 15, 'R': 15, 'S': 20, 'T': 10}
            
            for i, (col_letter, width) in enumerate(widths.items()):
                 if i < worksheet.max_column:
                    worksheet.column_dimensions[get_column_letter(i+1)].width = width

            money_cols = ["Số tiền trước Thuế", "Thuế 0%", "Thuế 5%", "Thuế 8%", "Thuế 10%", "Thuế khác", "Tiền thuế", "Số tiền sau"]
            money_col_indices = [export_df.columns.get_loc(c) + 1 for c in money_cols if c in export_df.columns]

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = border
                    cell.font = Font(name="Arial", size=10)
                    if cell.col_idx in money_col_indices:
                        cell.number_format = '#,##0'
                    
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    
        else:
            # === ACCOUNTING FORMAT (Existing) ===
            widths = {'A': 15, 'B': 15, 'C': 12, 'D': 15, 'E': 15, 'F': 20, 'G': 30, 'H': 18, 
                      'I': 15, 'J': 12, 'K': 10, 'L': 15, 'M': 18, 'N': 35}
            for col_letter, width in widths.items():
                worksheet.column_dimensions[col_letter].width = width

            # Format Data
            money_cols_idx = [9, 10, 12]  # I, J, L
            center_cols_idx = [1, 2, 3, 4, 5, 11]  # A-E, K
            
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    if isinstance(cell, openpyxl.cell.cell.MergedCell): continue
                    cell.border = border
                    cell.font = Font(name="Arial", size=10)
                    if cell.col_idx in money_cols_idx:
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif cell.col_idx in center_cols_idx:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(vertical="center", wrap_text=True)
            
            # Merge cells for multi-tax-rate invoices
            merge_by_file_cols = [9, 10, 12]  # I, J, L
            
            # First: Merge Team column by Team value (column A = 1)
            if len(export_df) > 0:
                start_row = 2
                current_team = worksheet.cell(row=2, column=1).value
                
                for excel_row in range(3, worksheet.max_row + 2):
                    if excel_row > worksheet.max_row:
                        cell_value = None
                    else:
                        cell_value = worksheet.cell(row=excel_row, column=1).value
                    
                    if cell_value != current_team:
                        end_row = excel_row - 1
                        if end_row > start_row:
                            worksheet.merge_cells(f"A{start_row}:A{end_row}")
                            top_cell = worksheet.cell(row=start_row, column=1)
                            top_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        
                        start_row = excel_row
                        current_team = cell_value
        
        # NOTE: Money columns (H, I, K) are NO LONGER merged
        # Each row shows its own tax rate and amount for clarity

    output.seek(0)
    
    st.download_button(
        label="💾 Tải file Excel kết quả",
        data=output,
        file_name="hoadon_tonghop.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        width='stretch'
    )

else:
    # === STEP 1: REQUIRED INPUTS ===
    st.markdown("### 📝 Bước 1: Thông tin bắt buộc")
    
    col1, col2 = st.columns(2)
    with col1:
        team_input = st.text_input("Team *", placeholder="Ví dụ: Team A, Team B...")
    with col2:
        employee_input = st.text_input("Tên nhân viên *", placeholder="Ví dụ: Nguyễn Văn A...")
    
    # === STEP 2: OPTIONAL CLASSIFICATION ===
    st.markdown("### 🏷️ Bước 2: Phân loại (Tùy chọn)")
    
    col_cat1, col_cat2 = st.columns(2)
    with col_cat1:
        category_select = st.selectbox("Chọn phân loại:", CATEGORY_OPTIONS)
    with col_cat2:
        custom_category = ""
        if category_select == "Khác (Nhập tay)":
            custom_category = st.text_input("Nhập phân loại tùy chỉnh:")

    # Report Type Selection
    report_type = st.radio("Chọn loại báo cáo đầu ra:", ["Kế toán", "Kinh doanh"], horizontal=True)

    
    st.divider()
    
    # === STEP 3: FILE UPLOAD ===
    st.markdown("### 📂 Bước 3: Tải hóa đơn (PDF)")
    
    # Check if required inputs are filled
    can_upload = bool(team_input.strip()) and bool(employee_input.strip())
    
    if not can_upload:
        st.warning("⚠️ Vui lòng nhập **Team** và **Tên nhân viên** trước khi tải file!")
    
    uploaded_files = st.file_uploader(
        "Kéo thả hoặc chọn nhiều file PDF vào đây", 
        type="pdf", 
        accept_multiple_files=True,
        disabled=not can_upload
    )

    if uploaded_files:
        st.divider()
        st.markdown("### ⚙️ Bước 4: Xử lý dữ liệu")
        st.write(f"Đã chọn **{len(uploaded_files)}** file.")
        
        if st.button("🚀 Bắt đầu trích xuất dữ liệu", type="primary"):
            logger.info(f"--- ACTION: Team={team_input}, Employee={employee_input} started processing {len(uploaded_files)} files ---")
            
            progress_bar = st.progress(0)
            status_box = st.empty()
            
            all_rows = []
            all_validations = []
            
            for i, uploaded_file in enumerate(uploaded_files):
                status_box.info(f"⏳ Đang xử lý: **{uploaded_file.name}** ({i+1}/{len(uploaded_files)})")
                progress_bar.progress((i + 1) / len(uploaded_files))
                
                try:
                    data, line_items = extract_invoice_data(uploaded_file, filename=uploaded_file.name)
                    uploaded_file.seek(0)
                    
                    # Validate extracted data
                    issues = validate_invoice_data(data)
                    all_validations.append({"file": uploaded_file.name, "issues": issues})
                    
                    # Determine classification
                    if category_select == "Khác (Nhập tay)" and custom_category.strip():
                        final_category = custom_category.strip()
                    elif category_select == "Tự động nhận diện":
                        # First check if OCR already set a classification
                        if data.get("Phân loại") and data.get("Phân loại") != "Khác":
                            final_category = data.get("Phân loại")
                        elif line_items:
                            all_item_names = " ".join([item.get("name", "") for item in line_items])
                            final_category = classify_content(all_item_names, data.get("Đơn vị bán", ""))
                        else:
                            final_category = classify_content("", data.get("Đơn vị bán", ""))
                    else:
                        final_category = category_select
                    
                    # Determine tax rate(s)
                    tax_rates = []
                    for rate in ["0%", "5%", "8%", "10%"]:
                        col_name = f"Thuế {rate}"
                        if data.get(col_name) and data.get(col_name) != "":
                            tax_rates.append(rate)
                    
                    if data.get("Thuế khác"):
                        tax_rates.append("Khác")
                    
                    if not tax_rates:
                        tax_rates = ["N/A"]
                    
                    # Create row(s) for this invoice
                    base_row = {
                        "Team": team_input.strip(),
                        "Số hóa đơn": data.get("Số hóa đơn", ""),
                        "Ngày hóa đơn": data.get("Ngày hóa đơn", ""),
                        "Mã số thuế bên bán": data.get("Mã số thuế", ""),
                        "Số ký hiệu": data.get("Ký hiệu", ""),
                        "Mã tra cứu": data.get("Mã tra cứu", ""),
                        "Link tra cứu": data.get("Link lấy hóa đơn", "") or data.get("Mã tra cứu", ""),
                        "Phân loại": final_category,
                        "Số tiền trước VAT": data.get("Số tiền trước Thuế", ""),
                        "Tổng tiền sau thuế": data.get("Số tiền sau", ""),
                        "Tên nhân viên": employee_input.strip(),
                        "Tên file": uploaded_file.name
                    }
                    
                    # Handle multi-rate invoices
                    # Helper function to parse money string
                    def parse_money_str(s):
                        if not s or pd.isna(s):
                            return 0
                        s = str(s).strip()
                        # Robust decimal handling: check order of separators
                        if '.' in s and ',' in s:
                            if s.rfind(',') > s.rfind('.'):
                                s = s[:s.rfind(',')]  # Cut decimal part
                            else:
                                s = s[:s.rfind('.')]
                        elif re.search(r'[,.]\\d{2}$', s) and not re.search(r'[,.]\\d{3}$', s):
                            s = s[:-3]  # Remove 2-digit decimal suffix
                        s = s.replace(',', '').replace('.', '')
                        try:
                            return int(s)
                        except:
                            return 0
                    
                    # Helper function to calculate amounts per tax rate
                    def calc_amounts_for_rate(vat_amount, rate_str):
                        """Calculate before-VAT and total from VAT amount and rate"""
                        vat_val = parse_money_str(vat_amount)
                        rate_map = {"0%": 0, "5%": 0.05, "8%": 0.08, "10%": 0.10}
                        rate = rate_map.get(rate_str, 0)
                        
                        if vat_val and rate > 0:
                            before_vat = int(round(vat_val / rate))
                            total = before_vat + vat_val
                            return before_vat, vat_val, total
                        elif vat_val:
                            return 0, vat_val, vat_val
                        return 0, 0, 0
                    
                    if report_type == "Kinh doanh":
                        # === BUSINESS FORMAT LOGIC (Wide) ===
                        business_row = {
                            "Team": team_input.strip(),
                            "Tên nhân viên": employee_input.strip(),
                            "Tên file": uploaded_file.name,
                            "Ngày hóa đơn": data.get("Ngày hóa đơn", ""),
                            "Số hóa đơn": data.get("Số hóa đơn", ""),
                            "Đơn vị bán": data.get("Đơn vị bán", ""),
                            "Phân loại": final_category,
                            "Số tiền trước Thuế": data.get("Số tiền trước Thuế", ""),
                            "Thuế 0%": data.get("Thuế 0%", ""),
                            "Thuế 5%": data.get("Thuế 5%", ""),
                            "Thuế 8%": data.get("Thuế 8%", ""),
                            "Thuế 10%": data.get("Thuế 10%", ""),
                            "Thuế khác": data.get("Thuế khác", ""),
                            "Tiền thuế": data.get("Tiền thuế", ""),
                            "Số tiền sau": data.get("Số tiền sau", ""),
                            "Link lấy hóa đơn": data.get("Link lấy hóa đơn", "") or data.get("Mã tra cứu", ""),
                            "Mã tra cứu": data.get("Mã tra cứu", ""),
                            "Mã số thuế": data.get("Mã số thuế", ""),
                            "Mã CQT": data.get("Mã CQT", ""),
                            "Ký hiệu": data.get("Ký hiệu", "")
                        }
                        all_rows.append(business_row)
                    else:
                        # === ACCOUNTING FORMAT LOGIC (Long) ===
                        if len(tax_rates) == 1:
                            # Single rate - simple case
                            rate = tax_rates[0]
                            if rate == "N/A":
                                base_row["VAT"] = data.get("Tiền thuế", "")
                                base_row["Thuế suất"] = ""
                                # Keep original totals for N/A
                            else:
                                vat_str = data.get(f"Thuế {rate}", data.get("Tiền thuế", ""))
                                base_row["VAT"] = vat_str
                                base_row["Thuế suất"] = rate
                                # ONLY calculate if extracted values are MISSING
                                # DO NOT overwrite already-extracted values!
                                if not base_row.get("Số tiền trước VAT") or not str(base_row.get("Số tiền trước VAT")).strip():
                                    before_vat, vat_val, total = calc_amounts_for_rate(vat_str, rate)
                                    if before_vat:
                                        base_row["Số tiền trước VAT"] = before_vat
                                    if total:
                                        base_row["Tổng tiền sau thuế"] = total
                            all_rows.append(base_row)
                        else:
                            # Multiple rates - create multiple rows with calculated amounts
                            for rate in tax_rates:
                                row = base_row.copy()
                                if rate == "Khác":
                                    row["VAT"] = data.get("Thuế khác", "")
                                    row["Thuế suất"] = "Khác"
                                else:
                                    vat_str = data.get(f"Thuế {rate}", "")
                                    before_vat, vat_val, total = calc_amounts_for_rate(vat_str, rate)
                                    row["VAT"] = vat_val if vat_val else vat_str
                                    row["Thuế suất"] = rate
                                    if before_vat:
                                        row["Số tiền trước VAT"] = before_vat
                                    if total:
                                        row["Tổng tiền sau thuế"] = total
                                all_rows.append(row)
                    
                except Exception as e:
                    logger.error(f"Error processing {uploaded_file.name}: {e}")
                    status_box.error(f"Lỗi khi xử lý {uploaded_file.name}")
                    all_validations.append({
                        "file": uploaded_file.name, 
                        "issues": [{"field": "File", "severity": "error", "message": f"Lỗi xử lý: {str(e)}"}]
                    })
            
            status_box.success("✅ Đã xử lý xong tất cả!")
            logger.info(f"--- COMPLETION: Team={team_input}, Employee={employee_input} finished processing ---")
            
            # Create DataFrame with appropriate columns
            if report_type == "Kinh doanh":
                columns = [
                    "Team", "Tên nhân viên", "Tên file", "Ngày hóa đơn", "Số hóa đơn", 
                    "Đơn vị bán", "Phân loại", "Số tiền trước Thuế", 
                    "Thuế 0%", "Thuế 5%", "Thuế 8%", "Thuế 10%", "Thuế khác", 
                    "Tiền thuế", "Số tiền sau", "Link lấy hóa đơn", 
                    "Mã tra cứu", "Mã số thuế", "Mã CQT", "Ký hiệu"
                ]
            else:
                 columns = [
                    "Team", "Số hóa đơn", "Ngày hóa đơn", "Mã số thuế bên bán", 
                    "Số ký hiệu", "Mã tra cứu", "Link tra cứu", "Phân loại", 
                    "Số tiền trước VAT", "VAT", "Thuế suất", "Tổng tiền sau thuế",
                    "Tên nhân viên", "Tên file"
                ]

            df = pd.DataFrame(all_rows)
            for col in columns:
                if col not in df.columns:
                    df[col] = ""
            df = df[columns]
            
            # Convert money columns
            if report_type == "Kinh doanh":
                 money_columns = ["Số tiền trước Thuế", "Thuế 0%", "Thuế 5%", "Thuế 8%", "Thuế 10%", "Thuế khác", "Tiền thuế", "Số tiền sau"]
            else:
                 money_columns = ["Số tiền trước VAT", "VAT", "Tổng tiền sau thuế"]

            for col in money_columns:
                def convert_to_number(x):
                    if pd.isna(x) or x == '': return None
                    x_str = str(x).strip()
                    if re.search(r',\d{2}$', x_str):
                        x_str = x_str.replace('.', '').replace(',', '.')
                    else:
                        x_str = x_str.replace('.', '').replace(',', '')
                    try:
                        return round(float(x_str))
                    except:
                        return x
                if col in df.columns:
                    df[col] = df[col].apply(convert_to_number)
            
            # Save to session state
            st.session_state["report_type"] = report_type
            st.session_state["processed_df"] = df
            st.session_state["validation_results"] = all_validations
            st.session_state["processing_complete"] = True
            st.rerun()
    else:
        st.info("👆 Vui lòng tải file lên để tiếp tục.")
